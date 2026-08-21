from flask import Flask, Response, jsonify, render_template, request, send_file, stream_with_context

import io
import json
import os
import time
import threading
import urllib
import urllib.request
import urllib.error

from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '.env'))

from datetime import datetime, timedelta, timezone

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:

    import psycopg

    from psycopg.rows import dict_row

except ImportError:

    psycopg = None

    dict_row = None

try:

    from supabase import create_client as _create_supabase_client

except ImportError:

    _create_supabase_client = None

# ============================================================
# 專業工控冷鏈最佳實踐標準參數（後端固化寫死，現場免手動繁瑣設定）
# ============================================================
ALARM_COOLDOWN_MIN   = 30   # 🔴 警報冷卻重響時間：30 分鐘 (HACCP/ISO 22000 食品冷鏈標準)
WARNING_COOLDOWN_MIN = 60   # 🟡 警告冷卻重響時間：60 分鐘 (1 小時低頻提醒，防群組洗版)
COMM_DEBOUNCE_SEC    = 180  # 🟡 通訊斷線防抖時間：180 秒 (連續 3 分鐘確認真斷線，過濾暫態雜訊)
BUZZER_SNOOZE_MIN    = 10   # 🔕 網頁消音後重響時間：10 分鐘 (消音後若未排除則 10 分鐘後重響)

app = Flask(__name__)

app.config['TEMPLATES_AUTO_RELOAD'] = True

@app.after_request

def add_cors_headers(response):

    response.headers['Access-Control-Allow-Origin'] = os.getenv('CORS_ORIGIN', '*')

    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'

    response.headers['Access-Control-Allow-Methods'] = 'GET,POST,OPTIONS'

    return response

DATABASE_URL = os.getenv('DATABASE_URL', '').strip()

if not DATABASE_URL:
    raise RuntimeError('DATABASE_URL is required (TimescaleDB/Postgres) — set it in .env')

# 設備控制命令佇列走 Supabase（不是本地 TimescaleDB）：
# 唯一真正碰得到現場 Modbus 的是 GCP 上的 gw1_supabase_collector.py，
# 它透過既有的 W610 連線輪詢 Supabase 的 device_commands 表來執行寫入。
SUPABASE_URL = os.getenv('SUPABASE_URL', '').strip()
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_ROLE_KEY', '').strip()
_supabase_client = None

def get_supabase():
    global _supabase_client
    if _supabase_client is None:
        if not _create_supabase_client:
            raise RuntimeError('supabase 套件未安裝，請先執行: pip install supabase')
        if not SUPABASE_URL or not SUPABASE_KEY:
            raise RuntimeError('缺少 SUPABASE_URL 或 SUPABASE_SERVICE_ROLE_KEY，請檢查 .env')
        _supabase_client = _create_supabase_client(SUPABASE_URL, SUPABASE_KEY)
    return _supabase_client

TZ_TW_APP = timezone(timedelta(hours=8))

REALTIME_LOCK = threading.Lock()

REALTIME_PAYLOAD = {}

# ── 主機運轉時數累積（壓縮機電流超過各 channel 自訂閾值 且 製冷中 才計時）──
RUNTIME_CURRENT_THRESHOLD_DEFAULT = float(os.getenv('RUNTIME_CURRENT_THRESHOLD', '3.0'))

RUNTIME_LOCK = threading.Lock()

RUNTIME_STATE = {}  # {channel: {'hours': float, 'last_ts': datetime}}

RUNTIME_STATE_LOADED = False

def _load_runtime_state():
    """啟動時從 DB 讀回每個 channel 最後一筆 runtime_hours，避免重啟歸零"""
    global RUNTIME_STATE_LOADED
    if RUNTIME_STATE_LOADED:
        return
    try:
        with get_pg() as conn:
            rows = conn.execute('''
                SELECT DISTINCT ON (channel) channel, runtime_hours
                FROM temperatures
                ORDER BY channel, timestamp DESC, id DESC
            ''').fetchall()
        for row in rows:
            RUNTIME_STATE[row['channel']] = {
                'hours': row['runtime_hours'] or 0.0,
                'last_ts': None
            }
    except Exception:
        pass
    RUNTIME_STATE_LOADED = True

def _update_runtime_hours(items):
    """依壓縮機電流是否超過各 channel 自訂閾值，累積每個 channel 的主機運轉時數（單位：小時）"""
    _load_runtime_state()
    now = datetime.now(TZ_TW_APP)
    alarm_map = _alarm_settings_map()
    with RUNTIME_LOCK:
        for item in items:
            channel = item.get('channel')
            if not channel:
                continue
            current = item.get('compressor_current')
            is_cooling = bool(item.get('cooling_status'))
            ch_setting = alarm_map.get(channel) or {}
            threshold = ch_setting.get('current_threshold')
            if threshold is None:
                threshold = RUNTIME_CURRENT_THRESHOLD_DEFAULT
            is_running = is_cooling and current is not None and float(current) > float(threshold)

            state = RUNTIME_STATE.setdefault(channel, {'hours': 0.0, 'last_ts': None})

            if is_running and state['last_ts'] is not None:
                elapsed_hours = (now - state['last_ts']).total_seconds() / 3600.0
                if 0 < elapsed_hours < 1:  # 忽略斷線重連造成的異常大跳動
                    state['hours'] += elapsed_hours

            state['last_ts'] = now
            item['runtime_hours'] = round(state['hours'], 2)

def get_pg():

    if not psycopg:

        raise RuntimeError('psycopg is required when DATABASE_URL is set')

    return psycopg.connect(DATABASE_URL, row_factory=dict_row)

def _fmt_ts(value):

    if isinstance(value, datetime):

        return value.astimezone(TZ_TW_APP).strftime('%Y-%m-%d %H:%M:%S')

    return value

def _payload_to_realtime(payload):

    readings = payload.get('readings', payload)

    if isinstance(readings, dict):

        items = [

            {'channel': ch, **info}

            for ch, info in readings.items()

            if isinstance(info, dict)

        ]

    elif isinstance(readings, list):

        items = readings

    else:

        raise ValueError('readings must be an object or array')

    timestamp = payload.get('timestamp') or datetime.now(TZ_TW_APP).strftime('%Y-%m-%d %H:%M:%S')

    alarm_map = _alarm_settings_map()

    result = {}

    for item in items:

        channel = item.get('channel')

        value = item.get('value')

        if not channel or value is None:

            continue

        value = float(value)

        ch_setting = alarm_map.get(channel, {})

        hi = ch_setting.get('hi')

        lo = ch_setting.get('lo')

        alarm_enabled = ch_setting.get('alarm_enabled')
        alarm_enabled = True if alarm_enabled is None else bool(alarm_enabled)

        result[channel] = {

            **item,

            'channel': channel,

            'name': item.get('name') or ch_setting.get('name') or channel,

            'value': value,

            'timestamp': item.get('timestamp') or timestamp,

            'hi': hi,

            'lo': lo,

            'in_alarm': alarm_enabled and ((hi is not None and value > hi) or (lo is not None and value < lo)),

            'status': item.get('status', 'NORMAL')

        }

    return result

def _alarm_settings_map():

    with get_pg() as conn:

        rows = conn.execute('SELECT * FROM alarm_settings ORDER BY channel').fetchall()

        return {r['channel']: dict(r) for r in rows}

def _latest_temperatures_payload():

    with REALTIME_LOCK:

        if REALTIME_PAYLOAD:

            return dict(REALTIME_PAYLOAD)

    with get_pg() as conn:

        rows = conn.execute('''

            SELECT DISTINCT ON (channel)

                channel, name, value, timestamp, status, runtime_hours

            FROM temperatures

            ORDER BY channel, timestamp DESC, id DESC

        ''').fetchall()

    alarm_map = _alarm_settings_map()

    result = {}

    for row in rows:

        ch = row['channel']

        value = row['value']

        ch_setting = alarm_map.get(ch, {})

        hi = ch_setting.get('hi')

        lo = ch_setting.get('lo')

        alarm_enabled = ch_setting.get('alarm_enabled')
        alarm_enabled = True if alarm_enabled is None else bool(alarm_enabled)

        status = row.get('status') or 'NORMAL'

        result[ch] = {

            'channel': ch,

            'name': row['name'],

            'value': value,

            'timestamp': _fmt_ts(row['timestamp']),

            'hi': hi,

            'lo': lo,

            'in_alarm': alarm_enabled and ((hi is not None and value > hi) or (lo is not None and value < lo)),

            'status': status,

            'runtime_hours': row.get('runtime_hours', 0.0)

        }

    return result

def _normalize_dt(val):
    if not val:
        return datetime.now(TZ_TW_APP)
    if isinstance(val, datetime):
        if val.tzinfo is None:
            return val.replace(tzinfo=TZ_TW_APP)
        return val.astimezone(TZ_TW_APP)
    if isinstance(val, str):
        s = val.strip().replace('T', ' ')
        try:
            if '+' in s or s.endswith('Z'):
                dt = datetime.fromisoformat(s.replace('Z', '+00:00'))
                return dt.astimezone(TZ_TW_APP)
            dt = datetime.strptime(s.split('.')[0], '%Y-%m-%d %H:%M:%S')
            return dt.replace(tzinfo=TZ_TW_APP)
        except Exception:
            return datetime.now(TZ_TW_APP)
    return datetime.now(TZ_TW_APP)

def _save_temperature_payload(payload):
    readings = payload.get('readings', payload)
    if isinstance(readings, dict):
        readings = [
            {'channel': ch, **info}
            for ch, info in readings.items()
            if isinstance(info, dict)
        ]
    if not isinstance(readings, list):
        raise ValueError('readings must be an object or array')

    now = datetime.now(TZ_TW_APP)
    ts_dt = _normalize_dt(payload.get('timestamp'))
    timestamp_str = ts_dt.strftime('%Y-%m-%d %H:%M:%S')

    saved = 0
    with get_pg() as conn:
        with conn.cursor() as cursor:
            for item in readings:
                channel = item.get('channel')
                name = item.get('name') or channel
                value = item.get('value')
                if not channel or value is None:
                    continue

                runtime_hours = float(item.get('runtime_hours', 0.0))
                ctrl_temp = float(item.get('control_temperature', value)) if (item.get('control_temperature') is not None or value is not None) else None
                coil_temp = float(item['coil_temperature']) if item.get('coil_temperature') is not None else None
                compressor_current = float(item['compressor_current']) if item.get('compressor_current') is not None else None
                high_pressure = float(item['high_pressure']) if item.get('high_pressure') is not None else None
                low_pressure = float(item['low_pressure']) if item.get('low_pressure') is not None else None
                set_temp = float(item['control_temperature_set']) if item.get('control_temperature_set') is not None else None

                cursor.execute('''
                    INSERT INTO temperatures (
                        timestamp, channel, name, value, status, runtime_hours,
                        control_temp, coil_temp, compressor_current, high_pressure, low_pressure, set_temp
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    ts_dt, channel, name, float(value), item.get('status', 'NORMAL'), runtime_hours,
                    ctrl_temp, coil_temp, compressor_current, high_pressure, low_pressure, set_temp
                ))
                saved += 1

                pw = item.get('power')
                if pw and isinstance(pw, dict):
                    v = pw.get('voltage_ll_avg') or pw.get('voltage_avg') or pw.get('v')
                    a = pw.get('current_avg') or pw.get('a')
                    kw = pw.get('kw') or (pw.get('power_total', 0) / 1000.0 if pw.get('power_total') is not None else None)
                    pf = pw.get('power_factor') or pw.get('pf')
                    kwh = pw.get('kwh') or (pw.get('energy_total', 0) / 1000.0 if pw.get('energy_total') is not None else None)
                    cursor.execute('''
                        INSERT INTO power_readings (timestamp, channel, v, a, kw, pf, kwh)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ''', (
                        ts_dt, channel,
                        float(v) if v is not None else None,
                        float(a) if a is not None else None,
                        float(kw) if kw is not None else None,
                        float(pf) if pf is not None else None,
                        float(kwh) if kwh is not None else None
                    ))

    return saved, timestamp_str

def _query_temp_range(date_from_str: str, date_to_str: str, channel: str = None):
    """
    溫度查詢：依時間範圍（與 channel，可選）查詢 temperatures。
    回傳 list[dict]，依 timestamp 升序排序。
    """
    dt_from = _normalize_dt(date_from_str)
    dt_to = _normalize_dt(date_to_str)
    q = '''
        SELECT id, timestamp, channel, name, value, status,
               COALESCE(control_temp, value) AS control_temp,
               coil_temp, compressor_current, high_pressure, low_pressure,
               set_temp, COALESCE(runtime_hours, 0.0) AS runtime_hours
        FROM temperatures
        WHERE timestamp BETWEEN %s AND %s
    '''
    params = [dt_from, dt_to]
    if channel:
        q += ' AND channel = %s'
        params.append(channel)
    q += ' ORDER BY timestamp'
    with get_pg() as conn:
        rows = conn.execute(q, params).fetchall()
    return [{**dict(r), 'timestamp': _fmt_ts(r['timestamp'])} for r in rows]

# ============================================================
# 12 通道預設設定
# ============================================================
DEFAULT_CHANNELS = [
    ('ch01', '1F 冷凍庫 A', -15.0, None),
    ('ch02', '1F 冷凍庫 B', -15.0, None),
    ('ch03', '1F 冷凍庫 C', -15.0, None),
    ('ch04', '1F 冷凍庫 D', -15.0, None),
    ('ch05', '1F 冷凍庫 E', -15.0, None),
    ('ch06', '1F 緩衝庫 A', 10.0, None),
    ('ch07', '1F 碼頭區 A', 15.0, None),
    ('ch08', '3F 急速庫 A', -15.0, None),
    ('ch09', '3F 急速庫 B', -15.0, None),
    ('ch10', '3F 半成品庫 A', 8.0, None),
    ('ch11', '3F 半成品庫 B', 8.0, None),
    ('ch12', '3F 冷藏庫 A', 8.0, None),
    ('ch13', '1F 集合式電錶', None, None),
    ('ch14', '3F 集合式電錶', None, None),
]

# D/E/F 庫：僅感溫棒，無壓縮機（iot627 資料只含控制溫度，不含運轉電流）
TEMP_ONLY_ROOMS = {'c', 'd', 'e'}

def init_db():
    with get_pg() as conn:
        with conn.cursor() as cursor:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS temperatures (
                    id BIGSERIAL PRIMARY KEY,
                    timestamp TIMESTAMPTZ NOT NULL,
                    channel TEXT NOT NULL,
                    name TEXT,
                    value DOUBLE PRECISION NOT NULL,
                    status TEXT DEFAULT 'NORMAL',
                    runtime_hours DOUBLE PRECISION DEFAULT 0.0,
                    control_temp DOUBLE PRECISION,
                    coil_temp DOUBLE PRECISION,
                    compressor_current DOUBLE PRECISION,
                    high_pressure DOUBLE PRECISION,
                    low_pressure DOUBLE PRECISION,
                    set_temp DOUBLE PRECISION
                )
            ''')
            for col, col_type in [
                ('runtime_hours', 'DOUBLE PRECISION DEFAULT 0.0'),
                ('control_temp', 'DOUBLE PRECISION'),
                ('coil_temp', 'DOUBLE PRECISION'),
                ('compressor_current', 'DOUBLE PRECISION'),
                ('high_pressure', 'DOUBLE PRECISION'),
                ('low_pressure', 'DOUBLE PRECISION'),
                ('set_temp', 'DOUBLE PRECISION'),
            ]:
                try:
                    cursor.execute(f'ALTER TABLE temperatures ADD COLUMN IF NOT EXISTS {col} {col_type}')
                except Exception:
                    pass

            cursor.execute('CREATE INDEX IF NOT EXISTS idx_temperatures_channel_time ON temperatures (channel, timestamp DESC)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_temperatures_time ON temperatures (timestamp DESC)')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS alarm_settings (
                    channel TEXT PRIMARY KEY,
                    name TEXT,
                    hi DOUBLE PRECISION,
                    lo DOUBLE PRECISION,
                    delay INTEGER DEFAULT 0,

                    alarm_enabled INTEGER DEFAULT 1,

                    temp_offset DOUBLE PRECISION DEFAULT 0.0,

                    current_threshold DOUBLE PRECISION DEFAULT 0.5,

                    nfb_rated_current DOUBLE PRECISION

                )

            ''')

            cursor.execute('''

                ALTER TABLE alarm_settings ADD COLUMN IF NOT EXISTS current_threshold DOUBLE PRECISION DEFAULT 0.5

            ''')

            cursor.execute('''
                ALTER TABLE alarm_settings ADD COLUMN IF NOT EXISTS nfb_rated_current DOUBLE PRECISION
            ''')
            cursor.execute('''
                ALTER TABLE alarm_settings ADD COLUMN IF NOT EXISTS power_anomaly_threshold DOUBLE PRECISION DEFAULT 15.0
            ''')

            cursor.execute('''
                ALTER TABLE alarm_settings DROP COLUMN IF EXISTS rated_load_pct
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS alarm_history (
                    id BIGSERIAL PRIMARY KEY,
                    triggered_at TIMESTAMPTZ NOT NULL,
                    channel TEXT,
                    name TEXT,
                    value DOUBLE PRECISION,
                    alarm_type TEXT,
                    hi DOUBLE PRECISION,
                    lo DOUBLE PRECISION,
                    category TEXT DEFAULT 'ALARM',
                    alarm_message TEXT,
                    restored_at TIMESTAMPTZ,
                    duration_sec INTEGER,
                    status TEXT DEFAULT 'ACTIVE'
                )
            ''')
            for col, col_type in [
                ('category', "TEXT DEFAULT 'ALARM'"),
                ('alarm_message', 'TEXT'),
                ('restored_at', 'TIMESTAMPTZ'),
                ('duration_sec', 'INTEGER'),
                ('status', "TEXT DEFAULT 'ACTIVE'"),
            ]:
                try:
                    cursor.execute(f'ALTER TABLE alarm_history ADD COLUMN IF NOT EXISTS {col} {col_type}')
                except Exception:
                    pass

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS monitoring_logs (
                    id          BIGSERIAL PRIMARY KEY,
                    timestamp   TIMESTAMPTZ NOT NULL,
                    site_code   TEXT NOT NULL,
                    device_no   TEXT NOT NULL,
                    data_key    TEXT NOT NULL,
                    data_value  DOUBLE PRECISION
                )
            ''')

            cursor.execute('CREATE INDEX IF NOT EXISTS idx_monitoring_logs_timestamp ON monitoring_logs (timestamp)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_monitoring_logs_device_no_key ON monitoring_logs (device_no, data_key)')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS power_readings (
                    id        BIGSERIAL PRIMARY KEY,
                    timestamp TIMESTAMPTZ NOT NULL,
                    channel   TEXT,
                    v         DOUBLE PRECISION,
                    a         DOUBLE PRECISION,
                    kw        DOUBLE PRECISION,
                    pf        DOUBLE PRECISION,
                    kwh       DOUBLE PRECISION
                )
            ''')

            cursor.execute('CREATE INDEX IF NOT EXISTS idx_power_readings_channel_time ON power_readings (channel, timestamp DESC)')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS hourly_power_stats (
                    hour_timestamp TIMESTAMPTZ PRIMARY KEY,
                    ch13_kwh_delta DOUBLE PRECISION DEFAULT 0.0,
                    ch14_kwh_delta DOUBLE PRECISION DEFAULT 0.0,
                    total_kwh_delta DOUBLE PRECISION DEFAULT 0.0,
                    tariff_type TEXT DEFAULT 'off_peak',
                    is_anomaly BOOLEAN DEFAULT FALSE
                )
            ''')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_hourly_power_stats_time ON hourly_power_stats (hour_timestamp DESC)')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS daily_power_stats (
                    date DATE PRIMARY KEY,
                    total_kwh DOUBLE PRECISION DEFAULT 0.0,
                    peak_kwh DOUBLE PRECISION DEFAULT 0.0,
                    semi_peak_kwh DOUBLE PRECISION DEFAULT 0.0,
                    off_peak_kwh DOUBLE PRECISION DEFAULT 0.0
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS system_config (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                )
            ''')
            cursor.execute('''
                INSERT INTO system_config (key, value) VALUES
                ('push_cooldown_min', '10'),
                ('buzzer_snooze_min', '10')
                ON CONFLICT (key) DO NOTHING
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS room_alarm_settings (
                    room_id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    channels TEXT NOT NULL,
                    hi DOUBLE PRECISION,
                    lo DOUBLE PRECISION,
                    delay INTEGER DEFAULT 10,
                    alarm_enabled INTEGER DEFAULT 1,
                    temp_offset DOUBLE PRECISION DEFAULT 0.0
                )
            ''')

            DEFAULT_ROOMS = [
                ('room1', '1F 冷凍庫', 'ch01,ch02,ch03,ch04,ch05', -15.0, -40.0, 10, 1, 0.0),
                ('room2', '1F 緩衝庫', 'ch06', 10.0, -40.0, 60, 0, 0.0),
                ('room3', '1F 碼頭區', 'ch07', 15.0, -40.0, 60, 0, 0.0),
                ('room4', '3F 急速庫', 'ch08,ch09', -15.0, -40.0, 60, 0, 0.0),
                ('room5', '3F 半成品冷凍庫', 'ch10,ch11', 8.0, -40.0, 60, 0, 0.0),
                ('room6', '3F 冷藏庫', 'ch12', 8.0, -40.0, 60, 0, 0.0),
            ]
            for r_id, r_name, chs, hi, lo, delay, enabled, offset in DEFAULT_ROOMS:
                cursor.execute('''
                    INSERT INTO room_alarm_settings (room_id, name, channels, hi, lo, delay, alarm_enabled, temp_offset)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (room_id) DO UPDATE
                    SET name = EXCLUDED.name, channels = EXCLUDED.channels
                ''', (r_id, r_name, chs, hi, lo, delay, enabled, offset))

            for ch, name, hi, lo in DEFAULT_CHANNELS:
                cursor.execute('''
                    INSERT INTO alarm_settings (channel, name, hi, lo)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (channel) DO NOTHING
                ''', (ch, name, hi, lo))

            for ch, name, _, _ in DEFAULT_CHANNELS:
                cursor.execute("UPDATE alarm_settings SET name = %s WHERE channel = %s", (name, ch))

def _get_all_running_hours():
    hours = {}
    try:
        with get_pg() as conn:
            with conn.cursor() as cursor:
                cursor.execute('''
                    SELECT site_code, device_no, COUNT(*)
                    FROM monitoring_logs AS m
                    WHERE data_key = 'current' AND (data_value >= 4.0 OR (data_value >= (
                        SELECT MIN(data_value)
                        FROM monitoring_logs
                        WHERE site_code = m.site_code AND device_no = m.device_no AND data_key = 'current'
                    ) + 1.0 AND data_value >= 1.5))
                    GROUP BY site_code, device_no
                ''')
                for r in cursor.fetchall():
                    hours[(r[0], r[1])] = round(r[2] / 60.0, 1)
    except Exception as e:
        print("Failed to query running hours from PG:", e)
    return hours

def _save_realtime_payload_to_logs(timestamp, payload):
    rows = []
    total_kwh = 0.0
    has_power = False
    
    ROOM_NAMES = {
        'a': 'A庫', 'b': 'B庫', 'c': 'D庫', 'd': 'E庫', 'e': 'F庫',
        'g': 'G庫', 'h': 'H庫', 'i1': 'I1庫', 'i2': 'I2庫', 'j': 'J庫', 'k': 'K庫'
    }
    
    for ch, d in payload.items():
        name = ROOM_NAMES.get(ch, ch.upper())
        val = d.get('value')
        if val is not None:
            rows.append((timestamp, name, 'SYSTEM', 'avg_temp', val))
            
        # Power
        pw = d.get('power')
        if pw:
            kw = pw.get('kw')
            kwh = pw.get('kwh')
            if kw is not None and kwh is not None:
                ROOM_PREFIXES = {
                    'a': 'A', 'b': 'B', 'c': 'D', 'd': 'E', 'e': 'F',
                    'g': 'G', 'h': 'H', 'i1': 'I1', 'i2': 'I2', 'j': 'J', 'k': 'K'
                }
                prefix = ROOM_PREFIXES.get(ch, ch.upper())
                meter_no = f"{prefix}-METER-01"
                rows.append((timestamp, name, meter_no, 'kw', kw))
                rows.append((timestamp, name, meter_no, 'kwh', kwh))
                total_kwh += kwh
                has_power = True
                
        # Units
        units = d.get('units')
        if units:
            for u in units:
                dev_id = u.get('id')
                temp = u.get('control_temperature')
                curr = u.get('compressor_current')
                if dev_id and temp is not None:
                    rows.append((timestamp, name, dev_id, 'temp_control', temp))
                    # D/E/F 庫為感溫棒（無壓縮機），不儲存運轉電流
                    if curr is not None and ch not in TEMP_ONLY_ROOMS:
                        rows.append((timestamp, name, dev_id, 'current', curr))
                    
    if has_power:
        rows.append((timestamp, '全廠', 'SYSTEM', 'total_kwh', round(total_kwh, 2)))
        
    if rows:
        try:
            with get_pg() as conn:
                with conn.cursor() as cursor:
                    cursor.executemany('''
                        INSERT INTO monitoring_logs (timestamp, site_code, device_no, data_key, data_value)
                        VALUES (%s, %s, %s, %s, %s)
                    ''', rows)
        except Exception as e:
            print("Postgres log save failed:", e)

def _record_temp_only_iot627_high_alarms(timestamp, payload):
    """Record D/E/F standalone IoT627 high-temperature alarms."""
    ROOM_NAMES = {
        'c': 'D庫',
        'd': 'E庫',
        'e': 'F庫',
    }
    rows = []
    for ch in TEMP_ONLY_ROOMS:
        d = payload.get(ch)
        if not isinstance(d, dict):
            continue
        hi = d.get('hi')
        if hi is None:
            continue
        room_name = ROOM_NAMES.get(ch, d.get('name') or ch.upper())
        units = d.get('units') or []
        iot_units = [u for u in units if isinstance(u, dict) and u.get('type') == 'iot627']
        if not iot_units:
            iot_units = [{'control_temperature': d.get('value')}]
        for unit in iot_units:
            value = unit.get('control_temperature')
            if value is None:
                continue
            value = float(value)
            if value > float(hi):
                rows.append((timestamp, ch, room_name, value, 'HIGH', hi, d.get('lo')))

    if not rows:
        return

    try:
        with get_pg() as conn:
            with conn.cursor() as cursor:
                for row in rows:
                    exists = cursor.execute('''
                        SELECT 1 FROM alarm_history
                        WHERE triggered_at = %s AND channel = %s AND alarm_type = %s
                        LIMIT 1
                    ''', (row[0], row[1], row[4])).fetchone()
                    if exists:
                        continue
                    cursor.execute('''
                        INSERT INTO alarm_history (triggered_at, channel, name, value, alarm_type, hi, lo)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ''', row)
    except Exception as e:
        print("Postgres alarm history save failed:", e)

def _start_mock_data_simulator():
    import random
    import threading
    import time
    import math
    from datetime import datetime
    
    base_temps = {
        'a': -18.0, 'b': -22.0, 'c': 4.0, 'd': 5.0, 'e': -15.0,
        'g': -20.0, 'h': 12.0, 'i1': 0.0, 'i2': 65.0, 'j': 24.0, 'k': 26.0
    }
    
    def run_simulator():
        last_saved_minute = None
        while True:
            payload = {}
            timestamp = datetime.now(TZ_TW_APP).strftime('%Y-%m-%d %H:%M:%S')
            hours_map = _get_all_running_hours()
            for ch in ['a', 'b', 'c', 'd', 'e', 'g', 'h', 'i1', 'i2', 'j', 'k']:
                base = base_temps[ch]
                val = base + random.uniform(-0.5, 0.5)
                base_temps[ch] = val
                
                ROOM_NAMES = {
                    'a': 'A庫', 'b': 'B庫', 'c': 'D庫', 'd': 'E庫', 'e': 'F庫',
                    'g': 'G庫', 'h': 'H庫', 'i1': 'I1庫', 'i2': 'I2庫', 'j': 'J庫', 'k': 'K庫'
                }
                name = ROOM_NAMES.get(ch, ch.upper())
                
                alarm_map = _alarm_settings_map()
                hi = alarm_map.get(ch, {}).get('hi')
                lo = alarm_map.get(ch, {}).get('lo')
                
                in_alarm = False
                status = 'NORMAL'
                if hi is not None and val > hi:
                    in_alarm = True
                    status = 'TRIGGERED'
                elif lo is not None and val < lo:
                    in_alarm = True
                    status = 'TRIGGERED'
                
                # Setup freezer units (iot627 and/or YB-D616-16DI)
                # D, E, F now have 1x iot627 (D-1, E-1, F-1) + 2x YB-D616-16DI
                units = []
                
                # Add iot627 module
                num_units = 3 if ch == 'i1' else (1 if ch in ('c', 'd', 'e') else 2)
                for idx in range(num_units):
                    ROOM_PREFIXES = {
                        'a': 'A', 'b': 'B', 'c': 'D', 'd': 'E', 'e': 'F',
                        'g': 'G', 'h': 'H', 'i1': 'I', 'i2': 'I', 'j': 'J', 'k': 'K'
                    }
                    prefix = ROOM_PREFIXES.get(ch, ch.upper())
                    dev_id = f"{prefix}-{idx + 1}"
                    if ch == 'i1':
                        dev_id = f"I-{idx + 1}"
                    elif ch == 'i2':
                        dev_id = f"I-{idx + 4}"
                    
                    setpoint_t = -18.0 if ch in ('a', 'b', 'e', 'g') else (4.0 if ch == 'c' else (5.0 if ch == 'd' else 65.0))
                    if ch == 'j': setpoint_t = 24.0
                    if ch == 'k': setpoint_t = 26.0

                    t_val = round(val - idx * 0.3, 1)

                    if ch in TEMP_ONLY_ROOMS:
                        # D/E/F 庫：感溫棒模式，無壓縮機，只記錄控制溫度
                        units.append({
                            'id': dev_id,
                            'type': 'iot627',
                            'control_temperature': t_val,
                            'coil_temperature': None,
                            'compressor_current': None,
                            'high_pressure': None,
                            'low_pressure': None,
                            'control_temperature_set': None,
                            'running_status': False,
                            'cooling_status': False,
                            'defrost_status': False,
                            'fan_status': False,
                            'eq_err': False,
                            'temp_err': in_alarm,
                            'total_running_hours': 0
                        })
                    else:
                        is_on = t_val > setpoint_t if base < 20.0 else t_val < setpoint_t
                        if ch in ('j', 'k'): is_on = t_val > setpoint_t

                        curr_val = round(12.5 + random.uniform(-0.8, 0.8), 2) if is_on else round(0.2 + random.uniform(-0.05, 0.05), 2)

                        db_hours = hours_map.get((name, dev_id), 0.0)

                        units.append({
                            'id': dev_id,
                            'type': 'iot627',
                            'control_temperature': t_val,
                            'coil_temperature': round(t_val - 12.0 - idx * 0.8 + random.uniform(-0.5, 0.5), 1),
                            'compressor_current': curr_val,
                            'high_pressure': round(1.45 - idx * 0.07 + random.uniform(-0.1, 0.1), 1),
                            'low_pressure': round(0.12 + idx * 0.03 + random.uniform(-0.02, 0.02), 1),
                            'control_temperature_set': setpoint_t,
                            'running_status': is_on,
                            'cooling_status': is_on,
                            'defrost_status': False,
                            'fan_status': True,
                            'eq_err': False,
                            'temp_err': in_alarm,
                            'total_running_hours': db_hours
                        })

                # Add YB-D616-16DI module if room has it
                has_di = ch in ('c', 'd', 'e', 'j', 'k')
                if has_di:
                    ROOM_PREFIXES = {
                        'a': 'A', 'b': 'B', 'c': 'D', 'd': 'E', 'e': 'F',
                        'g': 'G', 'h': 'H', 'i1': 'I', 'i2': 'I', 'j': 'J', 'k': 'K'
                    }
                    prefix = ROOM_PREFIXES.get(ch, ch.upper())
                    units.append({
                        'id': f"{prefix}-DI-01",
                        'type': 'YB-D616-16DI',
                        'connected': True,
                        'fan_01_running': True, 'fan_01_fault': False,
                        'fan_02_running': True, 'fan_02_fault': False,
                        'fan_03_running': False, 'fan_03_fault': False,
                        'fan_04_running': True, 'fan_04_fault': False,
                        'fan_05_running': True, 'fan_05_fault': False,
                        'fan_06_running': False, 'fan_06_fault': True,
                        'fan_07_running': True, 'fan_07_fault': False,
                        'fan_08_running': True, 'fan_08_fault': False,
                    })
                    units.append({
                        'id': f"{prefix}-DI-02",
                        'type': 'YB-D616-16DI',
                        'connected': True,
                        'fan_09_running': True, 'fan_09_fault': False,
                        'fan_10_running': True, 'fan_10_fault': False,
                        'fan_11_running': False, 'fan_11_fault': False,
                        'fan_12_running': True, 'fan_12_fault': False,
                    })

                idx_ch = ['a', 'b', 'c', 'd', 'e', 'g', 'h', 'i1', 'i2', 'j', 'k'].index(ch)
                rawV = 380.0 + random.uniform(-2.0, 2.0)
                rawA = 24.0 + random.uniform(-1.5, 1.5)
                rawKw = 13.5 + random.uniform(-0.8, 0.8)
                rawKwh = 1543.2 + (time.time() % 100) * 0.1
                rawPf = 0.88 + random.uniform(-0.02, 0.02)
                
                power = {
                    'v': round(rawV, 1),
                    'a': round(rawA, 1),
                    'kw': round(rawKw, 1),
                    'pf': round(rawPf, 1),
                    'kwh': round(rawKwh, 1),
                    'voltage_rs': round(rawV * (1 + math.sin(idx_ch) * 0.003), 1),
                    'voltage_st': round(rawV * (1 + math.cos(idx_ch) * 0.003), 1),
                    'voltage_rt': round(rawV * (1 - math.sin(idx_ch) * 0.003), 1),
                    'voltage_ll_avg': round(rawV, 1),
                    'current_r': round(rawA * (1 + math.sin(idx_ch) * 0.015), 1),
                    'current_s': round(rawA * (1 + math.cos(idx_ch) * 0.015), 1),
                    'current_t': round(rawA * (1 - math.sin(idx_ch) * 0.015), 1),
                    'current_avg': round(rawA, 1),
                    'power_total': round(rawKw * 1000.0, 1),
                    'energy_total': round(rawKwh * 1000.0, 1),
                    'power_factor': round(rawPf, 1)
                }
                
                payload[ch] = {
                    'channel': ch,
                    'name': alarm_map.get(ch, {}).get('name') or ch.upper(),
                    'value': round(val, 1),
                    'timestamp': timestamp,
                    'hi': hi,
                    'lo': lo,
                    'in_alarm': in_alarm,
                    'status': status,
                    'power': power,
                    'flags': {
                        'cooling': True,
                        'defrost': False,
                        'fan': True,
                        'eq_err': False,
                        'temp_err': in_alarm
                    },
                    'units': units
                }
            
            with REALTIME_LOCK:
                REALTIME_PAYLOAD.clear()
                REALTIME_PAYLOAD.update(payload)
            
            curr_min = timestamp[:16]
            if curr_min != last_saved_minute:
                _save_realtime_payload_to_logs(timestamp, payload)
                _record_temp_only_iot627_high_alarms(timestamp, payload)
                last_saved_minute = curr_min
                
            time.sleep(2.0)

    t = threading.Thread(target=run_simulator, daemon=True)
    t.start()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/remote')
def remote():
    client_ip = request.remote_addr
    forwarded = request.headers.get('X-Forwarded-For', '')
    if forwarded:
        client_ip = forwarded.split(',')[0].strip()
    
    allowed_ips = ['127.0.0.1', '61.222.3.117', '192.168.22.11']
    if client_ip not in allowed_ips and not client_ip.startswith('192.168.'):
        from flask import abort
        abort(403, description=f"拒絕存取：尚未授權此 IP ({client_ip}) 檢視遠端監控畫面。")
        
    return render_template('remote.html')

@app.route('/api/temperatures')
def temperatures():
    return jsonify(_latest_temperatures_payload())

@app.route('/api/temperatures', methods=['POST'])
def add_temperatures():
    try:
        payload = request.get_json(force=True) or {}
        raw_readings = payload.get('readings', payload)
        if isinstance(raw_readings, dict):
            items_for_runtime = [{'channel': ch, **info} for ch, info in raw_readings.items() if isinstance(info, dict)]
        elif isinstance(raw_readings, list):
            items_for_runtime = raw_readings
        else:
            items_for_runtime = []

        _update_runtime_hours(items_for_runtime)

        if isinstance(raw_readings, dict):
            for item in items_for_runtime:
                raw_readings[item['channel']]['runtime_hours'] = item.get('runtime_hours', 0.0)
        payload['readings'] = raw_readings

        realtime_payload = _payload_to_realtime(payload)

        with REALTIME_LOCK:
            REALTIME_PAYLOAD.clear()
            REALTIME_PAYLOAD.update(realtime_payload)

        if payload.get('realtime_only'):
            timestamp = payload.get('timestamp') or datetime.now(TZ_TW_APP).strftime('%Y-%m-%d %H:%M:%S')
            _record_temp_only_iot627_high_alarms(timestamp, realtime_payload)
            return jsonify({
                'status': 'ok',
                'saved': 0,
                'realtime': len(realtime_payload),
                'timestamp': payload.get('timestamp')
            })

        saved, timestamp = _save_temperature_payload(payload)
        _record_temp_only_iot627_high_alarms(timestamp, realtime_payload)

    except (TypeError, ValueError) as exc:
        return jsonify({'error': str(exc)}), 400
    return jsonify({'status': 'ok', 'saved': saved, 'timestamp': timestamp})

@app.route('/api/temperature_stream')
def temperature_stream():
    @stream_with_context
    def event_stream():
        last_payload = None
        while True:
            payload = _latest_temperatures_payload()
            encoded = json.dumps(payload, ensure_ascii=False)
            if encoded != last_payload:
                yield f"event: temperatures\ndata: {encoded}\n\n"
                last_payload = encoded
            else:
                yield ": keepalive\n\n"
            time.sleep(1)
    return Response(event_stream(), mimetype='text/event-stream')

@app.route('/api/alarm_settings', methods=['GET'])
def get_alarm_settings():
    result = {}
    for row in _alarm_settings_map().values():
        result[row['channel']] = {
            'channel': row['channel'],
            'name': row['name'],
            'hi': row['hi'],
            'lo': row['lo'],
            'delay': row.get('delay') or 0,
            'alarm_enabled': row.get('alarm_enabled') if row.get('alarm_enabled') is not None else 1,
            'temp_offset': row.get('temp_offset') or 0.0,
            'current_threshold': row.get('current_threshold') if row.get('current_threshold') is not None else 0.5,
            'nfb_rated_current': row.get('nfb_rated_current')
        }
    return jsonify(result)

@app.route('/api/alarm_settings', methods=['POST'])
def save_alarm_settings():
    data = request.json
    with get_pg() as conn:
        with conn.cursor() as cursor:
            for channel, setting in data.items():
                cursor.execute('''
                    UPDATE alarm_settings
                    SET hi=%s, lo=%s, delay=%s, alarm_enabled=%s, temp_offset=%s, current_threshold=%s, nfb_rated_current=%s
                    WHERE channel=%s
                ''', (
                    setting.get('hi'),
                    setting.get('lo'),
                    setting.get('delay', 0),
                    int(setting.get('alarm_enabled', 1)),
                    float(setting.get('temp_offset', 0)),
                    float(setting.get('current_threshold', 0.5)),
                    (float(setting['nfb_rated_current']) if setting.get('nfb_rated_current') not in (None, '') else None),
                    channel
                ))
    return jsonify({'status': 'ok'})

@app.route('/api/system_config', methods=['GET'])
def get_system_config():
    with get_pg() as conn:
        rows = conn.execute('SELECT key, value FROM system_config').fetchall()
        cfg = {r['key']: r['value'] for r in rows}
        return jsonify({
            'alarm_cooldown_min': int(float(cfg.get('alarm_cooldown_min', 30))),
            'warning_cooldown_min': int(float(cfg.get('warning_cooldown_min', 60))),
            'comm_debounce_sec': COMM_DEBOUNCE_SEC,
            'buzzer_snooze_min': int(float(cfg.get('buzzer_snooze_min', 10))),
            'display_resolution': cfg.get('display_resolution') or 'auto',
            'line_bot_enabled': int(float(cfg.get('line_bot_enabled', 0))),
            'line_channel_token': cfg.get('line_channel_token', ''),
            'line_target_id': cfg.get('line_target_id', '')
        })

@app.route('/api/system_config', methods=['POST'])
def save_system_config():
    data = request.json or {}
    keys = ['line_bot_enabled', 'line_channel_token', 'line_target_id', 'alarm_cooldown_min', 'warning_cooldown_min', 'buzzer_snooze_min', 'display_resolution']
    with get_pg() as conn:
        with conn.cursor() as cursor:
            for k in keys:
                if k in data and data[k] is not None:
                    val_str = str(data[k]).strip()
                    cursor.execute('''
                        INSERT INTO system_config (key, value) VALUES (%s, %s)
                        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
                    ''', (k, val_str))
        conn.commit()
    return jsonify({'status': 'ok'})

@app.route('/api/line_test', methods=['POST'])
def test_line_push():
    data = request.json or {}
    token = data.get('token', '').strip()
    target_id = data.get('target_id', '').strip()

    if not token or not target_id:
        return jsonify({'status': 'error', 'message': '請填寫 Channel Access Token 與 Target ID (User ID / Group ID)'}), 400

    test_msg = (
        "🔔【裕珍皇 · LINE Bot 測試連線】\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "恭喜！系統已成功連線至 LINE Messaging API 官方帳號。\n"
        "後續若發生冷鏈溫度警報、L212/L216 設備異常或通訊警告，將以此官方帳號自動推播通知。\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"⏰ 測試時間: {datetime.now(TZ_TW_APP).strftime('%Y-%m-%d %H:%M:%S')}"
    )

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }
    payload = {
        'to': target_id,
        'messages': [{'type': 'text', 'text': test_msg}]
    }
    try:
        req = urllib.request.Request(
            'https://api.line.me/v2/bot/message/push',
            data=json.dumps(payload).encode('utf-8'),
            headers=headers,
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            if resp.status == 200:
                return jsonify({'status': 'ok', 'message': '測試訊息已成功推播至 LINE！請檢查手機 LINE 訊息。'})
            return jsonify({'status': 'error', 'message': f'LINE API 回應狀態碼: {resp.status}'}), 500
    except urllib.error.HTTPError as e:
        err_body = e.read().decode('utf-8', errors='ignore')
        return jsonify({'status': 'error', 'message': f'LINE API 錯誤 ({e.code}): {err_body}'}), 400
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'連線失敗: {str(e)}'}), 500

def send_line_bot_flex(title, alt_text, header_bg, rows, footer_text=None):
    """
    發送金煜專案同款高質感 Flex Message（Bubble 視覺化彩色卡片）
    """
    try:
        with get_pg() as conn:
            cfg_rows = conn.execute("SELECT key, value FROM system_config WHERE key LIKE 'line_%'").fetchall()
            cfg = {r['key']: r['value'] for r in cfg_rows}
    except Exception:
        return False, "無法讀取 system_config"

    enabled = cfg.get('line_bot_enabled') == '1'
    token = cfg.get('line_channel_token', '').strip()
    target_id = cfg.get('line_target_id', '').strip()

    if not enabled or not token or not target_id:
        return False, "LINE Bot 未啟用或設定不完整"

    now_str = datetime.now(TZ_TW_APP).strftime('%Y-%m-%d %H:%M:%S')

    body_contents = []
    for item in rows:
        label = item[0]
        val = item[1]
        is_colored = item[2] if len(item) > 2 else False
        color = item[3] if len(item) > 3 else None

        row_box = {
            'type': 'box', 'layout': 'horizontal',
            'contents': [
                {'type': 'text', 'text': label, 'color': '#888888', 'size': 'sm', 'flex': 2},
                {'type': 'text', 'text': str(val), 'weight': 'bold', 'size': 'sm', 'flex': 5, 'wrap': True}
            ]
        }
        if is_colored and color:
            row_box['contents'][1]['color'] = color
        body_contents.append(row_box)

    body_contents.append({'type': 'separator'})
    body_contents.append({
        'type': 'text',
        'text': footer_text or '冷鏈監控系統持續守護中 🛡️',
        'color': '#888888', 'size': 'xs', 'wrap': True
    })

    flex_bubble = {
        'type': 'bubble', 'size': 'mega',
        'header': {
            'type': 'box', 'layout': 'vertical',
            'backgroundColor': header_bg,
            'contents': [
                {'type': 'text', 'text': title, 'color': '#FFFFFF', 'weight': 'bold', 'size': 'lg'},
                {'type': 'text', 'text': now_str, 'color': '#FFFFFF80', 'size': 'sm'}
            ]
        },
        'body': {
            'type': 'box', 'layout': 'vertical', 'spacing': 'md',
            'contents': body_contents
        }
    }

    payload = {
        'to': target_id,
        'messages': [{
            'type': 'flex',
            'altText': alt_text,
            'contents': flex_bubble
        }]
    }

    try:
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        req = urllib.request.Request(
            'https://api.line.me/v2/bot/message/push',
            data=json.dumps(payload).encode('utf-8'),
            headers=headers,
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            return resp.status == 200, "推播成功"
    except Exception as e:
        return False, f"LINE 推播失敗: {str(e)}"

def send_line_bot_notification(title, fields, is_alarm=False):
    """純文字備用通知介面"""
    rows = []
    for label, val in fields:
        rows.append((label, val, False, None))
    header_bg = '#FF3B30' if is_alarm else '#D97706'
    return send_line_bot_flex(title, title, header_bg, rows)

def push_gateway_rs485_disconnect(gw_id):
    """GW 斷線 ➔ 發報『設備 RS-485 通信異常』(金煜 Flex 卡片格式)"""
    gw_name = "1F 閘道器" if gw_id == "GW1" else "3F 閘道器"
    if gw_id == "GW1":
        affected_devices = [
            "1F 冷凍庫 A~E", "1F 緩衝庫", "1F 碼頭區", "1F 集合式電錶"
        ]
    else:
        affected_devices = [
            "3F 急速庫 (20HP/10HP)", "3F 半成品冷凍 A/B", "3F 冷藏庫", "3F 集合式電錶"
        ]

    rows = [
        ("斷線根因", f"{gw_name} (GW 連線中斷)", True, "#D97706"),
        ("警告類別", "設備 RS-485 通信異常", False, None),
        ("影響設備", "、".join(affected_devices), False, None)
    ]
    return send_line_bot_flex("⚠️ 設備 RS-485 通信異常", f"[警告] {gw_name} 設備 RS-485 通信異常", "#D97706", rows, "請檢查現場網關與 RS-485 通訊線路。")

def push_device_rs485_disconnect(device_name):
    """單一設備 RS-485 通信異常"""
    rows = [
        ("來源設備", device_name, True, "#D97706"),
        ("警告類別", "設備 RS-485 通信異常", False, None)
    ]
    return send_line_bot_flex("⚠️ 設備 RS-485 通信異常", f"[警告] {device_name} RS-485 通信異常", "#D97706", rows, "請檢查該設備從站接線與電源狀態。")

def push_temperature_alarm(room_name, alarm_type, val, threshold):
    """🔴 庫溫超標警報通知 (金煜同款 Flex Message 卡片)"""
    is_high = (alarm_type == 'HIGH')
    color = '#FF3B30' if is_high else '#007AFF'
    label = '高溫超標' if is_high else '低溫超標'
    limit = f'上限 {threshold:.1f}°C' if is_high else f'下限 {threshold:.1f}°C'
    alt_text = f'[警報] {room_name} {val:.1f}°C {label}'

    rows = [
        ("庫別名稱", room_name, False, None),
        ("目前溫度", f"{val:.1f}°C", True, color),
        ("警報類型", label, True, color),
        ("設定值", limit, False, None)
    ]
    return send_line_bot_flex("🚨 溫度警報通知", alt_text, color, rows, "請立即確認現場冷鏈設備狀態。")

def push_equipment_error_alarm(device_name, error_type):
    """🟠 L212 / 🟡 L216 設備異常通知（統一暖橘紅設計顏色，比高溫純紅降階）"""
    color = '#F97316'  # 統一暖橘紅色 (層次低於高溫純紅 #FF3B30，高於通信黃 #D97706)
    err_badge = f'{error_type} 設備異常'
    alt_text = f'[異常] {device_name} {err_badge}'

    rows = [
        ("來源設備", device_name, False, None),
        ("異常代碼", err_badge, True, color)
    ]
    return send_line_bot_flex("🚨 設備異常通知", alt_text, color, rows, "請立即指派現場技術人員檢修。")

def push_recovery_notice(target_name, event_type_desc, restored_devices=None):
    """🟢 溫度恢復正常通知 (金煜同款 Flex Message 卡片)"""
    color = '#34C759'
    alt_text = f'[恢復] {target_name} 溫度已恢復正常'

    rows = [
        ("庫別名稱", target_name, False, None),
        ("狀態說明", event_type_desc, True, color)
    ]
    if restored_devices:
        rows.append(("恢復設備", "、".join(restored_devices), False, None))
    return send_line_bot_flex("🟢 溫度恢復正常通知", alt_text, color, rows, "設備已恢復正常運作。")

_ROOM_ALARM_TRACKER = {}
_GW_ALARM_TRACKER = {}

def _evaluate_room_alarms_and_push(realtime_payload):
    """
    依據各「庫別」設定（hi/lo/delay/offset）與在線通道平均庫溫判定警報，
    並執行 LINE 官方帳號 Bot 主動推播與冷卻時間管理。
    """
    global _ROOM_ALARM_TRACKER, _GW_ALARM_TRACKER
    if not realtime_payload:
        return

    now_ts = time.time()
    now_dt = datetime.now(TZ_TW_APP)
    tw_now_str = now_dt.strftime('%Y-%m-%d %H:%M:%S')

    try:
        with get_pg() as conn:
            room_rows = conn.execute('SELECT room_id, name, channels, hi, lo, delay, alarm_enabled, temp_offset FROM room_alarm_settings').fetchall()
            cfg_rows = conn.execute("SELECT key, value FROM system_config").fetchall()
            cfg = {r['key']: r['value'] for r in cfg_rows}
    except Exception as e:
        print(f"Error fetching alarm settings in evaluator: {e}")
        return

    alarm_cooldown_min = int(float(cfg.get('alarm_cooldown_min', 30)))
    alarm_cooldown_sec = alarm_cooldown_min * 60

    for r in room_rows:
        room_id = r['room_id']
        r_name = r['name']
        chs = [c.strip() for c in (r['channels'] or '').split(',') if c.strip()]
        hi = r['hi']
        lo = r['lo']
        delay_min = int(r.get('delay') if r.get('delay') is not None else 10)
        delay_sec = delay_min * 60
        alarm_enabled = bool(r.get('alarm_enabled', 1))
        temp_offset = float(r.get('temp_offset') or 0.0)

        if room_id not in _ROOM_ALARM_TRACKER:
            _ROOM_ALARM_TRACKER[room_id] = {
                'pending_alarm_start': None,
                'pending_alarm_type': None,
                'is_alarm_active': False,
                'active_alarm_type': None,
                'active_alarm_val': None,
                'active_history_id': None,
                'last_push_time': 0
            }
        state = _ROOM_ALARM_TRACKER[room_id]

        # 計算在線通道平均庫溫
        vals = []
        for ch in chs:
            d = realtime_payload.get(ch)
            if d and isinstance(d, dict) and d.get('value') is not None:
                # 檢查時間戳是否為 45 秒內在線
                ts_str = d.get('timestamp')
                is_fresh = True
                if ts_str:
                    try:
                        d_ts = datetime.fromisoformat(ts_str.replace(' ', 'T')).timestamp()
                        if now_ts - d_ts > 45:
                            is_fresh = False
                    except Exception:
                        pass
                if is_fresh:
                    vals.append(float(d['value']))

        # 檢查該庫別目前在 DB 中是否有未解除的警報記錄
        db_active = False
        try:
            with get_pg() as conn:
                row = conn.execute("""
                    SELECT id, triggered_at FROM alarm_history
                    WHERE channel = %s AND (status = 'ACTIVE' OR restored_at IS NULL)
                    LIMIT 1
                """, (room_id,)).fetchone()
                if row:
                    db_active = True
        except Exception:
            pass

        if not vals or not alarm_enabled:
            # 庫別離線或停用警報 ➔ 若先前有警報則立即自動復歸
            if state['is_alarm_active'] or db_active:
                state['is_alarm_active'] = False
                state['active_history_id'] = None
                try:
                    with get_pg() as conn:
                        with conn.cursor() as cur:
                            cur.execute("""
                                UPDATE alarm_history
                                SET status = 'CLEARED', restored_at = %s,
                                    duration_sec = ROUND(EXTRACT(EPOCH FROM (%s - triggered_at)))
                                WHERE channel = %s AND (status = 'ACTIVE' OR restored_at IS NULL)
                            """, (tw_now_str, tw_now_str, room_id))
                        conn.commit()
                except Exception as e:
                    print(f"Error clearing alarm_history: {e}")
            continue

        avg_temp = round((sum(vals) / len(vals)) + temp_offset, 1)

        # 判定高低溫
        triggered_type = None
        threshold_val = None
        if hi is not None and avg_temp > hi:
            triggered_type = 'HIGH'
            threshold_val = hi
        elif lo is not None and avg_temp < lo:
            triggered_type = 'LOW'
            threshold_val = lo

        if triggered_type:
            if state['pending_alarm_type'] != triggered_type:
                state['pending_alarm_start'] = now_ts
                state['pending_alarm_type'] = triggered_type

            time_abnormal = now_ts - (state['pending_alarm_start'] or now_ts)
            if time_abnormal >= delay_sec:
                if not state['is_alarm_active'] and not db_active:
                    # 首次觸發警報
                    state['is_alarm_active'] = True
                    state['active_alarm_type'] = triggered_type
                    state['active_alarm_val'] = avg_temp
                    state['last_push_time'] = now_ts

                    # 寫入 alarm_history
                    msg = f"{'庫溫過高警報' if triggered_type == 'HIGH' else '庫溫過低警報'} ({avg_temp:.1f}°C, 門檻: {threshold_val:.1f}°C)"
                    try:
                        with get_pg() as conn:
                            with conn.cursor() as cur:
                                cur.execute("""
                                    INSERT INTO alarm_history (triggered_at, channel, name, value, alarm_type, hi, lo, category, alarm_message, status)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'ALARM', %s, 'ACTIVE')
                                    RETURNING id
                                """, (tw_now_str, room_id, r_name, avg_temp, triggered_type, hi, lo, msg))
                                row = cur.fetchone()
                                if row:
                                    state['active_history_id'] = row['id']
                            conn.commit()
                    except Exception as e:
                        print(f"Error inserting alarm_history: {e}")

                    # 發送 LINE Bot 推播
                    print(f"🚨 [ALARM TRIGGERED] {r_name} {triggered_type} ({avg_temp}°C > {threshold_val}°C) -> Pushing to LINE...")
                    push_temperature_alarm(r_name, triggered_type, avg_temp, threshold_val)
                else:
                    # 警報持續中，檢查冷卻時間再次發報
                    state['is_alarm_active'] = True
                    if now_ts - state['last_push_time'] >= alarm_cooldown_sec:
                        state['last_push_time'] = now_ts
                        print(f"🚨 [ALARM COOLDOWN EXPIRED] Re-pushing {r_name} {triggered_type} to LINE...")
                        push_temperature_alarm(r_name, triggered_type, avg_temp, threshold_val)
        else:
            # 庫溫正常 ➔ 若先前處於警報狀態，立即執行復歸！
            state['pending_alarm_start'] = None
            state['pending_alarm_type'] = None

            if state['is_alarm_active'] or db_active:
                state['is_alarm_active'] = False
                state['active_history_id'] = None

                try:
                    with get_pg() as conn:
                        with conn.cursor() as cur:
                            cur.execute("""
                                UPDATE alarm_history
                                SET status = 'CLEARED', restored_at = %s,
                                    duration_sec = ROUND(EXTRACT(EPOCH FROM (%s - triggered_at)))
                                WHERE channel = %s AND (status = 'ACTIVE' OR restored_at IS NULL)
                            """, (tw_now_str, tw_now_str, room_id))
                        conn.commit()
                except Exception as e:
                    print(f"Error clearing alarm_history: {e}")

                print(f"🟢 [ALARM CLEARED] {r_name} restored to normal ({avg_temp}°C) -> Sending recovery notice to LINE...")
                push_recovery_notice(r_name, f"庫溫已恢復正常範圍 (目前均溫: {avg_temp:.1f}°C)")

def _start_alarm_monitor_worker():
    """啟動背景常駐警報評估與推播監控線程"""
    import threading
    def _worker():
        while True:
            try:
                with REALTIME_LOCK:
                    current_payload = dict(REALTIME_PAYLOAD)
                if current_payload:
                    _evaluate_room_alarms_and_push(current_payload)
            except Exception as e:
                print(f"Alarm monitor worker error: {e}")
            time.sleep(3.0)
    t = threading.Thread(target=_worker, daemon=True)
    t.start()

_start_alarm_monitor_worker()

@app.route('/api/room_alarm_settings', methods=['GET'])
def get_room_alarm_settings():
    with get_pg() as conn:
        rows = conn.execute('SELECT room_id, name, channels, hi, lo, delay, alarm_enabled, temp_offset FROM room_alarm_settings ORDER BY room_id').fetchall()
        result = {}
        for r in rows:
            result[r['room_id']] = {
                'room_id': r['room_id'],
                'name': r['name'],
                'channels': [c.strip() for c in r['channels'].split(',') if c.strip()],
                'hi': r['hi'],
                'lo': r['lo'],
                'delay': r.get('delay') if r.get('delay') is not None else 10,
                'alarm_enabled': r.get('alarm_enabled') if r.get('alarm_enabled') is not None else 1,
                'temp_offset': r.get('temp_offset') or 0.0
            }
        return jsonify(result)

@app.route('/api/room_alarm_settings', methods=['POST'])
def save_room_alarm_settings():
    data = request.json or {}
    with get_pg() as conn:
        with conn.cursor() as cursor:
            for room_id, setting in data.items():
                hi = setting.get('hi')
                lo = setting.get('lo')
                delay = int(setting.get('delay', 10))
                enabled = int(setting.get('alarm_enabled', 1))
                offset = float(setting.get('temp_offset', 0.0))
                
                cursor.execute('''
                    UPDATE room_alarm_settings
                    SET hi=%s, lo=%s, delay=%s, alarm_enabled=%s, temp_offset=%s
                    WHERE room_id=%s
                ''', (hi, lo, delay, enabled, offset, room_id))

                # 同步更新該庫別對應之所有通道 alarm_settings
                channels_list = setting.get('channels', [])
                if not channels_list:
                    r_row = conn.execute('SELECT channels FROM room_alarm_settings WHERE room_id=%s', (room_id,)).fetchone()
                    if r_row and r_row['channels']:
                        channels_list = [c.strip() for c in r_row['channels'].split(',') if c.strip()]

                for ch in channels_list:
                    cursor.execute('''
                        UPDATE alarm_settings
                        SET hi=%s, lo=%s, delay=%s, alarm_enabled=%s, temp_offset=%s
                        WHERE channel=%s
                    ''', (hi, lo, delay, enabled, offset, ch))
        conn.commit()

    # 儲存後立即重新評估警報狀態（若調高門檻則立即復歸解除！）
    try:
        with REALTIME_LOCK:
            current_payload = dict(REALTIME_PAYLOAD)
        if current_payload:
            _evaluate_room_alarms_and_push(current_payload)
    except Exception as e:
        print(f"Post-save alarm evaluation error: {e}")

    return jsonify({'status': 'ok'})

# ============================================================

# 警報歷史

# ============================================================

@app.route('/api/alarm_history', methods=['GET'])
def alarm_history():
    channel    = request.args.get('channel', 'all')
    category   = request.args.get('category', 'all')
    alarm_type = request.args.get('alarm_type', 'all')
    status     = request.args.get('status', 'all')
    date_from  = request.args.get('from', None)
    date_to    = request.args.get('to',   None)
    limit      = int(request.args.get('limit', 300))

    ROOM_CHANNEL_MAP = {
        'room1': ['ch01', 'ch02', 'ch03', 'ch04', 'ch05', 'room1'],
        'room2': ['ch06', 'room2'],
        'room3': ['ch07', 'room3'],
        'room4': ['ch08', 'ch09', 'room4'],
        'room5': ['ch10', 'ch11', 'room5'],
        'room6': ['ch12', 'room6'],
    }

    conditions, params = [], []
    if channel != 'all':
        if channel in ROOM_CHANNEL_MAP:
            chs = ROOM_CHANNEL_MAP[channel]
            placeholders = ', '.join(['%s'] * len(chs))
            conditions.append(f'channel IN ({placeholders})')
            params.extend(chs)
        else:
            conditions.append('channel = %s')
            params.append(channel)

    if category != 'all':
        conditions.append('category = %s')
        params.append(category)

    if alarm_type != 'all':
        conditions.append('alarm_type = %s')
        params.append(alarm_type)

    if status == 'active':
        conditions.append("(status = 'ACTIVE' OR restored_at IS NULL)")
    elif status == 'cleared':
        conditions.append("(status = 'CLEARED' OR restored_at IS NOT NULL)")

    if date_from:
        conditions.append('triggered_at >= %s')
        params.append(date_from)

    if date_to:
        to_str = date_to if len(date_to) > 10 else date_to + ' 23:59:59'
        conditions.append('triggered_at <= %s')
        params.append(to_str)

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

    now_tw = datetime.now(TZ_TW_APP)
    month_start = now_tw.replace(day=1, hour=0, minute=0, second=0, microsecond=0).strftime('%Y-%m-%d %H:%M:%S')

    with get_pg() as conn:
        records = conn.execute(f'''
            SELECT id, triggered_at, channel, name, value, alarm_type, hi, lo,
                   category, alarm_message, restored_at, duration_sec, status
            FROM alarm_history {where}
            ORDER BY triggered_at DESC LIMIT %s
        ''', params + [limit]).fetchall()

        # 頂部三大本月卡片運算：當前未解除、本月高溫警報、本月設備異常
        try:
            active_row = conn.execute("SELECT COUNT(*) AS c FROM alarm_history WHERE status = 'ACTIVE' OR restored_at IS NULL").fetchone()
            active_cnt = active_row['c'] if active_row else 0
        except Exception:
            active_cnt = 0

        try:
            high_row = conn.execute("SELECT COUNT(*) AS c FROM alarm_history WHERE (alarm_type = 'HIGH' OR alarm_type = 'HIGH_TEMP') AND triggered_at >= %s", [month_start]).fetchone()
            month_high_cnt = high_row['c'] if high_row else 0
        except Exception:
            month_high_cnt = 0

        try:
            equip_row = conn.execute("SELECT COUNT(*) AS c FROM alarm_history WHERE (alarm_type IN ('EQUIP_STOP_L212', 'EQUIP_ERR_L216', 'EQUIP', 'L212', 'L216') OR category = 'EQUIP') AND triggered_at >= %s", [month_start]).fetchone()
            month_equip_cnt = equip_row['c'] if equip_row else 0
        except Exception:
            month_equip_cnt = 0

    formatted_records = []
    for r in records:
        d = dict(r)
        d['triggered_at'] = _fmt_ts(d['triggered_at'])
        d['restored_at'] = _fmt_ts(d['restored_at']) if d.get('restored_at') else None
        formatted_records.append(d)

    return jsonify({
        'records': formatted_records,
        'total': len(formatted_records),
        'stats': {
            'active_count': active_cnt,
            'month_high_temp_count': month_high_cnt,
            'month_equip_err_count': month_equip_cnt
        }
    })

@app.route('/api/alarm_history', methods=['POST'])
def add_alarm_history():
    data = request.json or {}
    tw_now = datetime.now(TZ_TW_APP).strftime('%Y-%m-%d %H:%M:%S')

    category = data.get('category', 'ALARM')
    alarm_type = data.get('alarm_type', 'HIGH')
    msg = data.get('alarm_message') or data.get('message') or f"{'高溫警報' if alarm_type=='HIGH' else '低溫警報'} ({data.get('value')}°C)"

    with get_pg() as conn:
        conn.execute('''
            INSERT INTO alarm_history (triggered_at, channel, name, value, alarm_type, hi, lo, category, alarm_message, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (tw_now, data.get('channel'), data.get('name'), data.get('value'),
              alarm_type, data.get('hi'), data.get('lo'), category, msg, 'ACTIVE'))

    return jsonify({'status': 'ok'})

@app.route('/api/alarm_history/export')
def export_alarm_history():
    channel    = request.args.get('channel', 'all')
    category   = request.args.get('category', 'all')
    alarm_type = request.args.get('alarm_type', 'all')
    status     = request.args.get('status', 'all')
    date_from  = request.args.get('from', None)
    date_to    = request.args.get('to',   None)

    ROOM_CHANNEL_MAP = {
        'room1': ['ch01', 'ch02', 'ch03', 'ch04', 'ch05', 'room1'],
        'room2': ['ch06', 'room2'],
        'room3': ['ch07', 'room3'],
        'room4': ['ch08', 'ch09', 'room4'],
        'room5': ['ch10', 'ch11', 'room5'],
        'room6': ['ch12', 'room6'],
    }

    conditions, params = [], []
    if channel != 'all':
        if channel in ROOM_CHANNEL_MAP:
            chs = ROOM_CHANNEL_MAP[channel]
            placeholders = ', '.join(['%s'] * len(chs))
            conditions.append(f'channel IN ({placeholders})')
            params.extend(chs)
        else:
            conditions.append('channel = %s')
            params.append(channel)
    if category != 'all':
        conditions.append('category = %s')
        params.append(category)
    if alarm_type != 'all':
        conditions.append('alarm_type = %s')
        params.append(alarm_type)
    if status == 'active':
        conditions.append("(status = 'ACTIVE' OR restored_at IS NULL)")
    elif status == 'cleared':
        conditions.append("(status = 'CLEARED' OR restored_at IS NOT NULL)")

    if date_from:
        conditions.append('triggered_at >= %s')
        params.append(date_from)
    if date_to:
        to_str = date_to if len(date_to) > 10 else date_to + ' 23:59:59'
        conditions.append('triggered_at <= %s')
        params.append(to_str)

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

    with get_pg() as conn:
        records = conn.execute(f'''
            SELECT id, triggered_at, channel, name, value, alarm_type, hi, lo,
                   category, alarm_message, restored_at, duration_sec, status
            FROM alarm_history {where}
            ORDER BY triggered_at DESC LIMIT 2000
        ''', params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = '警報記錄'

    hdr_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='0C3B82')
    cell_border = Border(left=Side(style='thin', color='CBD5E1'),
                         right=Side(style='thin', color='CBD5E1'),
                         top=Side(style='thin', color='CBD5E1'),
                         bottom=Side(style='thin', color='CBD5E1'))

    headers = ['#', '觸發時間', '類別', '來源位置', '警報內容', '觸發數值', '門檻值', '狀態', '復歸時間', '持續時長']
    widths = [6, 20, 14, 20, 34, 12, 12, 12, 20, 16]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[chr(64+col)].width = w

    for r_idx, r in enumerate(records, 2):
        d = dict(r)
        trig_t = _fmt_ts(d['triggered_at'])
        rest_t = _fmt_ts(d['restored_at']) if d.get('restored_at') else '--'
        st_label = '警報中' if (d.get('status') == 'ACTIVE' or not d.get('restored_at')) else '已復歸'
        dur_label = f"{d.get('duration_sec', 0) // 60}分{d.get('duration_sec', 0) % 60}秒" if d.get('duration_sec') else ('發報中' if st_label == '警報中' else '--')
        msg = d.get('alarm_message') or (f"{'高溫警報' if d.get('alarm_type')=='HIGH' else '低溫警報'} ({d.get('value')}°C)")

        row_vals = [
            r_idx - 1, trig_t, d.get('category') or 'ALARM',
            d.get('name') or d.get('channel'), msg,
            d.get('value') if d.get('value') is not None else '--',
            d.get('hi') if d.get('alarm_type') == 'HIGH' else (d.get('lo') if d.get('lo') is not None else '--'),
            st_label, rest_t, dur_label
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(r_idx, col, val)
            c.border = cell_border
            c.alignment = Alignment(horizontal='center' if col in (1, 3, 6, 7, 8, 10) else 'left', vertical='center')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'警報歷史紀錄_{datetime.now(TZ_TW_APP).strftime("%Y%m%d_%H%M%S")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ============================================================

# 設備控制命令佇列
# 前端寫入 pending 命令 → Supabase device_commands 表
# → GCP 上的 gw1_supabase_collector.py 用既有的 W610 連線輪詢執行、回報結果
# set_temperature：對應 IoT627 40007 控制溫度設定
# set_power_state：對應 IoT627 40179 (A801 停控模式)，1=全開 0=全停（實測確認位址）
# ============================================================

SUPPORTED_COMMAND_TYPES = {'set_temperature', 'set_power_state'}

@app.route('/api/device_commands', methods=['POST'])
def create_device_command():
    data = request.json or {}
    channel = data.get('channel')
    command_type = data.get('command_type')
    value = data.get('value')

    if not channel:
        return jsonify({'error': 'channel is required'}), 400
    if command_type not in SUPPORTED_COMMAND_TYPES:
        return jsonify({'error': f'unsupported command_type: {command_type}'}), 400
    if command_type == 'set_temperature':
        try:
            value = float(value)
        except (TypeError, ValueError):
            return jsonify({'error': 'value must be a number for set_temperature'}), 400
        if not (-60 <= value <= 60):
            return jsonify({'error': 'value out of safe range (-60~60 degC)'}), 400
    elif command_type == 'set_power_state':
        try:
            value = float(value)
        except (TypeError, ValueError):
            return jsonify({'error': 'value must be 0 or 1 for set_power_state'}), 400
        if value not in (0, 1):
            return jsonify({'error': 'value must be 0 (停) or 1 (開) for set_power_state'}), 400

    resp = get_supabase().table('device_commands').insert({
        'channel': channel,
        'command_type': command_type,
        'value': value,
        'status': 'pending'
    }).execute()

    return jsonify({'status': 'ok', 'id': resp.data[0]['id']})

@app.route('/api/device_commands', methods=['GET'])
def list_device_commands():
    status = request.args.get('status')
    channel = request.args.get('channel')
    limit = int(request.args.get('limit', 50))

    q = get_supabase().table('device_commands').select('*')
    if status:
        q = q.eq('status', status)
    if channel:
        q = q.eq('channel', channel)
    resp = q.order('created_at', desc=True).limit(limit).execute()

    return jsonify(resp.data)

@app.route('/api/device_commands/<int:command_id>', methods=['GET'])
def get_device_command(command_id):
    resp = get_supabase().table('device_commands').select('*').eq('id', command_id).execute()
    if not resp.data:
        return jsonify({'error': 'not found'}), 404
    return jsonify(resp.data[0])

@app.route('/api/device_commands/<int:command_id>/complete', methods=['POST'])
def complete_device_command(command_id):
    data = request.json or {}
    success = bool(data.get('success'))
    error_message = data.get('error_message')

    get_supabase().table('device_commands').update({
        'status': 'success' if success else 'failed',
        'error_message': error_message,
        'executed_at': datetime.now(TZ_TW_APP).strftime('%Y-%m-%d %H:%M:%S')
    }).eq('id', command_id).execute()

    return jsonify({'status': 'ok'})

# ============================================================

# 圖表資料

# ============================================================

@app.route('/api/chart_data')
def chart_data():
    range_type = request.args.get('range', 'realtime')
    date_from  = request.args.get('from', None)
    date_to    = request.args.get('to',   None)
    now_tw = datetime.now(TZ_TW_APP)

    if range_type == 'custom' and date_from and date_to:
        from_str = date_from
        to_str   = date_to
    else:
        minutes_map = {'realtime': 60, '1h': 60, '4h': 240, '6h': 360, '24h': 1440}
        minutes = minutes_map.get(range_type, 60)
        from_str = (now_tw - timedelta(minutes=minutes)).strftime('%Y-%m-%d %H:%M:%S')
        to_str   = now_tw.strftime('%Y-%m-%d %H:%M:%S')

    rows = _query_temp_range(from_str, to_str)
    alarm_map = _alarm_settings_map()

    series = {}
    for row in rows:
        ch = row['channel']
        if ch not in series:
            series[ch] = {'name': row['name'] or ch, 'data': []}
        series[ch]['data'].append({
            't': row['timestamp'],
            'v': row['value'],
            'control_temp': row.get('control_temp', row['value']),
            'coil_temp': row.get('coil_temp'),
            'compressor_current': row.get('compressor_current'),
            'high_pressure': row.get('high_pressure'),
            'low_pressure': row.get('low_pressure'),
            'set_temp': row.get('set_temp'),
            'runtime_hours': row.get('runtime_hours', 0.0)
        })

    stats = {}
    for ch, s in series.items():
        vals  = [d['v'] for d in s['data'] if d.get('v') is not None]
        times = [d['t'] for d in s['data']]
        if vals:
            max_v = max(vals)
            min_v = min(vals)
            hi    = alarm_map.get(ch, {}).get('hi')
            lo    = alarm_map.get(ch, {}).get('lo')
            ch_alarm_enabled = alarm_map.get(ch, {}).get('alarm_enabled')
            ch_alarm_enabled = True if ch_alarm_enabled is None else bool(ch_alarm_enabled)
            stats[ch] = {
                'name':     s['name'],
                'max':      round(max_v, 1),
                'max_time': times[vals.index(max_v)],
                'min':      round(min_v, 1),
                'min_time': times[vals.index(min_v)],
                'avg':      round(sum(vals)/len(vals), 1),
                'count':    len(vals),
                'in_alarm': ch_alarm_enabled and any(
                    (hi is not None and v > hi) or (lo is not None and v < lo)
                    for v in vals
                ),
                'hi': hi,
                'lo': lo
            }

    return jsonify({'series': series, 'stats': stats, 'alarm_map': alarm_map})

@app.route('/api/power_chart_data')
def power_chart_data():
    channels   = request.args.getlist('ch')   # e.g. ?ch=ch13&ch=ch14 or ?ch=all
    field      = request.args.get('field', 'delta_kwh')   # delta_kwh / power_total / energy_total / etc
    range_type = request.args.get('range', '6h')
    date_from  = request.args.get('from', None)
    date_to    = request.args.get('to',   None)

    if not channels or 'all' in channels:
        channels = ['ch13', 'ch14']

    now_tw = datetime.now(TZ_TW_APP)
    if range_type == 'custom' and date_from and date_to:
        dt_from = _normalize_dt(date_from)
        dt_to   = _normalize_dt(date_to)
    else:
        minutes_map = {'6h': 360, '1d': 1440, '24h': 1440, 'day': 1440, '7d': 10080, 'week': 10080, '1h': 60, 'realtime': 60}
        minutes  = minutes_map.get(range_type, 360)
        dt_from = now_tw - timedelta(minutes=minutes)
        dt_to   = now_tw

    total_seconds = (dt_to - dt_from).total_seconds()
    if total_seconds <= 6 * 3600:
        bucket_sec = 300       # 5 分鐘
    elif total_seconds <= 24 * 3600:
        bucket_sec = 900       # 15 分鐘
    elif total_seconds <= 7 * 86400:
        bucket_sec = 3600      # 1 小時
    else:
        bucket_sec = 86400     # 1 天

    series = {}
    with get_pg() as conn:
        for ch in channels:
            rows = conn.execute(
                "SELECT timestamp, kw, kwh, v, a, pf FROM power_readings "
                "WHERE channel=%s AND timestamp BETWEEN %s AND %s ORDER BY timestamp ASC",
                (ch, dt_from, dt_to)
            ).fetchall()

            if field == 'delta_kwh':
                # 計算各時間切片的實質用電增量 (ΔkWh)
                buckets = {}
                for r in rows:
                    ts = r['timestamp']
                    if not ts.tzinfo:
                        ts = ts.replace(tzinfo=timezone.utc).astimezone(TZ_TW_APP)
                    else:
                        ts = ts.astimezone(TZ_TW_APP)
                    epoch = int(ts.timestamp())
                    b_epoch = (epoch // bucket_sec) * bucket_sec
                    if b_epoch not in buckets:
                        buckets[b_epoch] = {'min_kwh': r['kwh'], 'max_kwh': r['kwh'], 'ts': datetime.fromtimestamp(b_epoch, TZ_TW_APP)}
                    b = buckets[b_epoch]
                    if r['kwh'] is not None:
                        if b['min_kwh'] is None or r['kwh'] < b['min_kwh']:
                            b['min_kwh'] = r['kwh']
                        if b['max_kwh'] is None or r['kwh'] > b['max_kwh']:
                            b['max_kwh'] = r['kwh']

                sorted_epochs = sorted(buckets.keys())
                data_pts = []
                for idx, ep in enumerate(sorted_epochs):
                    b = buckets[ep]
                    delta = 0.0
                    if b['max_kwh'] is not None and b['min_kwh'] is not None and b['max_kwh'] >= b['min_kwh']:
                        delta = float(b['max_kwh'] - b['min_kwh'])
                    if delta == 0.0 and idx > 0:
                        prev_b = buckets[sorted_epochs[idx - 1]]
                        if b['max_kwh'] is not None and prev_b['max_kwh'] is not None:
                            diff = float(b['max_kwh'] - prev_b['max_kwh'])
                            if 0 <= diff < 500:
                                delta = diff
                    data_pts.append({
                        't': b['ts'].strftime('%Y-%m-%d %H:%M:%S'),
                        'v': round(delta, 2)
                    })
                series[ch] = data_pts
            else:
                col_map = {
                    'power_total': 'kw', 'kw': 'kw',
                    'energy_total': 'kwh', 'kwh': 'kwh',
                    'voltage_ll_avg': 'v', 'v': 'v',
                    'current_avg': 'a', 'a': 'a',
                    'power_factor': 'pf', 'pf': 'pf'
                }
                c_name = col_map.get(field, 'kw')
                series[ch] = [{
                    't': _fmt_ts(r['timestamp']),
                    'v': round(float(r[c_name]), 2) if r[c_name] is not None else None
                } for r in rows]

    return jsonify({'series': series, 'field': field, 'bucket_sec': bucket_sec})

# ── 台電時間電價尖離峰時段判斷 ────────────────────────────────────
OFF_PEAK_HOLIDAYS = [
    "01-01",  # 元旦
    "02-28",  # 二二八和平紀念日
    "04-04",  # 兒童節
    "04-05",  # 清明節
    "05-01",  # 勞動節
    "10-10",  # 國慶日
]

TAIPOWER_TARIFF_RULES = {
    "summer": {
        "weekday": [
            {"start_hour": 0, "end_hour": 9, "type": "off_peak"},
            {"start_hour": 9, "end_hour": 16, "type": "semi_peak"},
            {"start_hour": 16, "end_hour": 22, "type": "peak"},       # 16:00~22:00 夜尖峰
            {"start_hour": 22, "end_hour": 24, "type": "semi_peak"},
        ],
        "saturday": [
            {"start_hour": 0, "end_hour": 9, "type": "off_peak"},
            {"start_hour": 9, "end_hour": 24, "type": "semi_peak"},
        ],
        "sunday_and_holidays": [
            {"start_hour": 0, "end_hour": 24, "type": "off_peak"},
        ],
    },
    "non_summer": {
        "weekday": [
            {"start_hour": 0, "end_hour": 6, "type": "off_peak"},
            {"start_hour": 6, "end_hour": 16, "type": "semi_peak"},
            {"start_hour": 16, "end_hour": 22, "type": "peak"},
            {"start_hour": 22, "end_hour": 24, "type": "semi_peak"},
        ],
        "saturday": [
            {"start_hour": 0, "end_hour": 6, "type": "off_peak"},
            {"start_hour": 6, "end_hour": 24, "type": "semi_peak"},
        ],
        "sunday_and_holidays": [
            {"start_hour": 0, "end_hour": 24, "type": "off_peak"},
        ],
    },
}

def get_tariff_type(dt: datetime, is_high_voltage: bool = True) -> str:
    """判斷指定時間點的時間電價時段 (peak / semi_peak / off_peak)"""
    date_str = dt.strftime("%m-%d")
    time_hour = dt.hour
    weekday = dt.weekday()  # 0=週一, 5=週六, 6=週日

    if is_high_voltage:
        is_summer = "05-16" <= date_str <= "10-15"
    else:
        is_summer = "06-01" <= date_str <= "09-30"

    season_key = "summer" if is_summer else "non_summer"

    if weekday == 6 or date_str in OFF_PEAK_HOLIDAYS:
        day_type = "sunday_and_holidays"
    elif weekday == 5:
        day_type = "saturday"
    else:
        day_type = "weekday"

    rules = TAIPOWER_TARIFF_RULES.get(season_key, {}).get(day_type, [])
    for rule in rules:
        if rule["start_hour"] <= time_hour < rule["end_hour"]:
            return rule["type"]

    return "off_peak"

@app.route('/api/power_energy_stats')
def power_energy_stats():
    period = request.args.get('period', 'month')  # day / week / month / custom
    date_from_str = request.args.get('from', None)
    date_to_str = request.args.get('to', None)

    now_tw = datetime.now(TZ_TW_APP)
    today_start = now_tw.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # 本週起始 (週一 00:00)
    week_start = (now_tw - timedelta(days=now_tw.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = now_tw.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # 決定當前選定週期的起始與結束時間
    if period == 'day':
        dt_period_from = today_start
        dt_period_to = now_tw
    elif period == 'week':
        dt_period_from = week_start
        dt_period_to = now_tw
    elif period == 'custom' and date_from_str and date_to_str:
        dt_period_from = _normalize_dt(date_from_str)
        dt_period_to = _normalize_dt(date_to_str)
    else: # month (預設)
        dt_period_from = month_start
        dt_period_to = now_tw

    lookback_start = (now_tw - timedelta(days=35)).replace(hour=0, minute=0, second=0, microsecond=0)
    hourly_data = {}

    with get_pg() as conn:
        for ch in ['ch13', 'ch14']:
            rows = conn.execute(
                "SELECT timestamp, kwh, kw FROM power_readings "
                "WHERE channel=%s AND timestamp >= %s ORDER BY timestamp ASC",
                (ch, lookback_start)
            ).fetchall()
            if not rows:
                continue

            hourly_buckets = {}
            for r in rows:
                ts = r['timestamp']
                if not ts.tzinfo:
                    ts = ts.replace(tzinfo=timezone.utc).astimezone(TZ_TW_APP)
                else:
                    ts = ts.astimezone(TZ_TW_APP)
                bucket_key = (ts.strftime('%Y-%m-%d'), ts.hour)
                if bucket_key not in hourly_buckets:
                    hourly_buckets[bucket_key] = {'min_kwh': r['kwh'], 'max_kwh': r['kwh'], 'kw_sum': 0, 'kw_cnt': 0, 'dt': ts}
                b = hourly_buckets[bucket_key]
                if r['kwh'] is not None:
                    if b['min_kwh'] is None or r['kwh'] < b['min_kwh']:
                        b['min_kwh'] = r['kwh']
                    if b['max_kwh'] is None or r['kwh'] > b['max_kwh']:
                        b['max_kwh'] = r['kwh']
                if r['kw'] is not None:
                    b['kw_sum'] += float(r['kw'])
                    b['kw_cnt'] += 1

            sorted_keys = sorted(hourly_buckets.keys())
            for idx, k in enumerate(sorted_keys):
                b = hourly_buckets[k]
                dt_hour = b['dt'].replace(minute=0, second=0, microsecond=0)
                tariff = get_tariff_type(dt_hour)
                delta = 0.0
                if b['max_kwh'] is not None and b['min_kwh'] is not None and b['max_kwh'] >= b['min_kwh']:
                    delta = float(b['max_kwh'] - b['min_kwh'])
                if delta == 0.0 and idx > 0:
                    prev_k = sorted_keys[idx - 1]
                    prev_b = hourly_buckets[prev_k]
                    if b['max_kwh'] is not None and prev_b['max_kwh'] is not None:
                        diff = float(b['max_kwh'] - prev_b['max_kwh'])
                        if 0 <= diff < 1500:
                            delta = diff
                if delta == 0.0 and b['kw_cnt'] > 0:
                    avg_kw = b['kw_sum'] / b['kw_cnt']
                    if avg_kw > 0:
                        delta = min(avg_kw, 500.0)

                if k not in hourly_data:
                    hourly_data[k] = {'ch13': 0.0, 'ch14': 0.0, 'total': 0.0, 'tariff': tariff, 'dt': dt_hour}
                hourly_data[k][ch] = delta
                hourly_data[k]['total'] += delta

    # 統計本日、本週、當月用電
    today_kwh = 0.0
    week_kwh = 0.0
    month_kwh = 0.0

    # 選定週期內的統計 (尖峰、半尖峰、離峰)
    period_total_kwh = 0.0
    period_peak_kwh = 0.0
    period_semi_peak_kwh = 0.0
    period_off_peak_kwh = 0.0

    for (d_str, hr), h in hourly_data.items():
        dt_h = h['dt']
        tot = h['total']
        if dt_h >= today_start:
            today_kwh += tot
        if dt_h >= week_start:
            week_kwh += tot
        if dt_h >= month_start:
            month_kwh += tot

        if dt_period_from <= dt_h <= dt_period_to:
            period_total_kwh += tot
            if h['tariff'] == 'peak':
                period_peak_kwh += tot
            elif h['tariff'] == 'semi_peak':
                period_semi_peak_kwh += tot
            else:
                period_off_peak_kwh += tot

    if period_total_kwh > 0:
        peak_ratio = round((period_peak_kwh / period_total_kwh) * 100, 1)
        semi_peak_ratio = round((period_semi_peak_kwh / period_total_kwh) * 100, 1)
        off_peak_ratio = round((period_off_peak_kwh / period_total_kwh) * 100, 1)
    else:
        peak_ratio = 0.0
        semi_peak_ratio = 0.0
        off_peak_ratio = 0.0

    DEVICE_NAMES_MAP = {
        'ch01': '1F 冷凍庫 A', 'ch02': '1F 冷凍庫 B', 'ch03': '1F 冷凍庫 C',
        'ch04': '1F 冷凍庫 D', 'ch05': '1F 冷凍庫 E', 'ch06': '1F 緩衝庫 A',
        'ch07': '1F 碼頭區 A', 'ch08': '3F 急速庫 20HP', 'ch09': '3F 急速庫 10HP',
        'ch10': '3F 半成品庫 A', 'ch11': '3F 半成品庫 B', 'ch12': '3F 冷藏庫 A',
        'ch13': '1F 集合式電錶', 'ch14': '3F 集合式電錶'
    }

    # 12 台冷鏈設備即時工況與選定週期累積時數
    # 庫別定義
    ROOM_MAP = [
        {'id': 'room1', 'name': '1F 冷凍庫', 'channels': ['ch01', 'ch02', 'ch03', 'ch04', 'ch05'], 'color': '#1a5fa8'},
        {'id': 'room2', 'name': '1F 緩衝庫', 'channels': ['ch06'], 'color': '#16a085'},
        {'id': 'room3', 'name': '1F 碼頭區', 'channels': ['ch07'], 'color': '#27ae60'},
        {'id': 'room4', 'name': '3F 急速庫', 'channels': ['ch08', 'ch09'], 'color': '#c0392b'},
        {'id': 'room5', 'name': '3F 半成品冷凍庫', 'channels': ['ch10', 'ch11'], 'color': '#e67e22'},
        {'id': 'room6', 'name': '3F 冷藏庫', 'channels': ['ch12'], 'color': '#8e44ad'}
    ]

    realtime_data = _latest_temperatures_payload()
    alarm_map = _alarm_settings_map()

    # 計算週期內各設備運轉時數
    device_runtimes = []
    room_scores = {r['id']: 0.0 for r in ROOM_MAP}

    with get_pg() as conn:
        for r_def in ROOM_MAP:
            for ch in r_def['channels']:
                # 查詢該設備在該週期內的最早與最晚 runtime_hours 差值
                r_rows = conn.execute('''
                    SELECT runtime_hours FROM temperatures
                    WHERE channel = %s AND timestamp BETWEEN %s AND %s AND runtime_hours IS NOT NULL
                    ORDER BY timestamp ASC
                ''', (ch, dt_period_from, dt_period_to)).fetchall()

                r_hours = 0.0
                if r_rows and len(r_rows) >= 2:
                    h_start = float(r_rows[0]['runtime_hours'] or 0.0)
                    h_end = float(r_rows[-1]['runtime_hours'] or 0.0)
                    if h_end >= h_start:
                        r_hours = round(h_end - h_start, 1)
                elif r_rows and len(r_rows) == 1:
                    r_hours = 0.5

                cur_info = realtime_data.get(ch, {})
                comp_curr = cur_info.get('compressor_current')
                cooling = cur_info.get('cooling_status') or cur_info.get('flags', {}).get('cooling')
                defrost = cur_info.get('defrost_status') or cur_info.get('flags', {}).get('defrost')
                eq_err = cur_info.get('eq_err') or cur_info.get('flags', {}).get('eq_err')

                status = 'stop'
                if eq_err:
                    status = 'error'
                elif defrost:
                    status = 'defrost'
                elif cooling and comp_curr and float(comp_curr) > (alarm_map.get(ch, {}).get('current_threshold') or 0.5):
                    status = 'cooling'

                device_runtimes.append({
                    'ch': ch,
                    'room_id': r_def['id'],
                    'room_name': r_def['name'],
                    'name': DEVICE_NAMES_MAP.get(ch, ch),
                    'status': status,
                    'current': round(float(comp_curr), 1) if comp_curr is not None else 0.0,
                    'runtime_hours': r_hours
                })

                # 依電流與時數權重累計該庫能耗積分
                eff_curr = float(comp_curr) if comp_curr else (4.5 if '急速' in r_def['name'] else 2.5)
                room_scores[r_def['id']] += max(r_hours * eff_curr, 0.1)

    # 計算各庫圓餅圖佔比
    tot_score = sum(room_scores.values()) or 1.0
    actual_kwh_base = period_total_kwh if period_total_kwh > 0 else 100.0

    room_distribution = []
    for r_def in ROOM_MAP:
        pct = round((room_scores[r_def['id']] / tot_score) * 100, 1)
        kwh = round((pct / 100.0) * actual_kwh_base, 1)
        room_distribution.append({
            'id': r_def['id'],
            'name': r_def['name'],
            'kwh': kwh,
            'pct': pct,
            'color': r_def['color']
        })

    return jsonify({
        'status': 'ok',
        'period': period,
        'today_kwh': round(today_kwh, 1),
        'week_kwh': round(week_kwh, 1),
        'month_kwh': round(month_kwh, 1),
        'period_total_kwh': round(period_total_kwh, 1),
        'period_peak_kwh': round(period_peak_kwh, 1),
        'period_semi_peak_kwh': round(period_semi_peak_kwh, 1),
        'period_off_peak_kwh': round(period_off_peak_kwh, 1),
        'peak_ratio': peak_ratio,
        'semi_peak_ratio': semi_peak_ratio,
        'off_peak_ratio': off_peak_ratio,
        'dt_from': dt_period_from.strftime('%Y-%m-%d %H:%M'),
        'dt_to': dt_period_to.strftime('%Y-%m-%d %H:%M'),
        'room_distribution': room_distribution,
        'device_runtimes': device_runtimes
    })

# ============================================================
# 報表查詢與 Excel 匯出引擎 (冷鏈庫溫歷程紀錄表 & SPM-3 電表能源統計報表)
# ============================================================

REPORT_ROOM_DEFINITIONS = {
    'room1': {'name': '1F 冷凍庫', 'channels': ['ch01', 'ch02', 'ch03', 'ch04', 'ch05']},
    'room2': {'name': '1F 緩衝庫', 'channels': ['ch06']},
    'room3': {'name': '1F 碼頭區', 'channels': ['ch07']},
    'room4': {'name': '3F 急速庫', 'channels': ['ch08', 'ch09']},
    'room5': {'name': '3F 半成品冷凍庫', 'channels': ['ch10', 'ch11']},
    'room6': {'name': '3F 冷藏庫', 'channels': ['ch12']},
}

REPORT_METER_DEFINITIONS = {
    'all': {'name': '全廠電表資訊', 'channels': ['ch13', 'ch14']},
    'ch13': {'name': '1F 集合式電錶', 'channels': ['ch13']},
    'ch14': {'name': '3F 集合式電錶', 'channels': ['ch14']},
}

def _calculate_report_time_range(period, date_from_str=None, date_to_str=None):
    now_tw = datetime.now(TZ_TW_APP)
    today = now_tw.date()
    
    if period == 'day':
        if date_from_str:
            try:
                target_date = datetime.strptime(date_from_str[:10], '%Y-%m-%d').date()
            except Exception:
                target_date = today
        else:
            target_date = today
        d_from = datetime.combine(target_date, datetime.min.time(), tzinfo=TZ_TW_APP)
        d_to = datetime.combine(target_date, datetime.max.time(), tzinfo=TZ_TW_APP)
        label = f"{target_date.strftime('%Y-%m-%d')} (日報表)"
        file_suffix = target_date.strftime('%Y-%m-%d')
    elif period == 'week':
        if date_from_str:
            try:
                ref_date = datetime.strptime(date_from_str[:10], '%Y-%m-%d').date()
            except Exception:
                ref_date = today
        else:
            ref_date = today
        days_since_monday = ref_date.weekday()
        week_monday = ref_date - timedelta(days=days_since_monday)
        week_sunday = week_monday + timedelta(days=6)
        d_from = datetime.combine(week_monday, datetime.min.time(), tzinfo=TZ_TW_APP)
        d_to = datetime.combine(week_sunday, datetime.max.time(), tzinfo=TZ_TW_APP)
        label = f"{week_monday.strftime('%Y-%m-%d')} ~ {week_sunday.strftime('%Y-%m-%d')} (週報表)"
        file_suffix = f"W_{week_monday.strftime('%Y%m%d')}_{week_sunday.strftime('%Y%m%d')}"
    elif period == 'month':
        if date_from_str:
            try:
                ref_date = datetime.strptime(date_from_str[:10], '%Y-%m-%d').date()
            except Exception:
                ref_date = today
        else:
            ref_date = today
        first_day_month = ref_date.replace(day=1)
        import calendar
        _, last_day_num = calendar.monthrange(ref_date.year, ref_date.month)
        last_day_month = ref_date.replace(day=last_day_num)
        d_from = datetime.combine(first_day_month, datetime.min.time(), tzinfo=TZ_TW_APP)
        d_to = datetime.combine(last_day_month, datetime.max.time(), tzinfo=TZ_TW_APP)
        label = f"{first_day_month.strftime('%Y-%m')} (月報表)"
        file_suffix = first_day_month.strftime('%Y-%m')
    elif period == 'custom':
        if not date_from_str:
            date_from_str = today.strftime('%Y-%m-%d')
        if not date_to_str:
            date_to_str = today.strftime('%Y-%m-%d')
        dt_start = datetime.strptime(date_from_str[:10], '%Y-%m-%d').date()
        dt_end = datetime.strptime(date_to_str[:10], '%Y-%m-%d').date()
        d_from = datetime.combine(dt_start, datetime.min.time(), tzinfo=TZ_TW_APP)
        d_to = datetime.combine(dt_end, datetime.max.time(), tzinfo=TZ_TW_APP)
        label = f"{dt_start.strftime('%Y-%m-%d')} ~ {dt_end.strftime('%Y-%m-%d')} (自訂區間)"
        file_suffix = f"{dt_start.strftime('%Y%m%d')}_{dt_end.strftime('%Y%m%d')}"
    else:
        raise ValueError("無效的查詢週期")

    return d_from, d_to, label, file_suffix

def _generate_temperature_report_data(room_id, d_from, d_to, interval_min=5):
    r_def = REPORT_ROOM_DEFINITIONS.get(room_id)
    if not r_def:
        raise ValueError("未知的庫別代碼")
    
    room_name = r_def['name']
    chs = r_def['channels']
    
    hi, lo = None, None
    offset = 0.0
    with get_pg() as conn:
        r_set = conn.execute("SELECT hi, lo, temp_offset FROM room_alarm_settings WHERE room_id = %s", (room_id,)).fetchone()
        if r_set:
            hi = r_set['hi']
            lo = r_set['lo']
            offset = float(r_set.get('temp_offset') or 0.0)

        chs_placeholder = ', '.join(['%s'] * len(chs))
        rows = conn.execute(f"""
            SELECT timestamp, channel, name, value, control_temp, coil_temp, compressor_current
            FROM temperatures
            WHERE channel IN ({chs_placeholder}) AND timestamp >= %s AND timestamp <= %s
            ORDER BY timestamp ASC
        """, chs + [d_from, d_to]).fetchall()

    from collections import defaultdict
    buckets = defaultdict(lambda: {'ts_list': [], 'vals': []})
    
    for r in rows:
        ts = r['timestamp']
        if isinstance(ts, datetime):
            ts_tw = ts.astimezone(TZ_TW_APP)
        else:
            ts_tw = datetime.fromisoformat(str(ts).replace(' ', 'T')).replace(tzinfo=TZ_TW_APP)
        
        floored_min = (ts_tw.minute // interval_min) * interval_min
        bucket_time = ts_tw.replace(minute=floored_min, second=0, microsecond=0)
        b_key = bucket_time.strftime('%Y-%m-%d %H:%M')
        
        v = r['value'] if r['value'] is not None else r['control_temp']
        if v is not None:
            buckets[b_key]['ts_list'].append(ts_tw)
            buckets[b_key]['vals'].append(float(v) + offset)

    records = []
    curr = d_from.replace(minute=(d_from.minute // interval_min) * interval_min, second=0, microsecond=0)
    end_curr = d_to.replace(second=0, microsecond=0)
    
    all_avg_temps = []
    out_of_bounds_count = 0
    
    while curr <= end_curr:
        b_key = curr.strftime('%Y-%m-%d %H:%M')
        b_data = buckets.get(b_key)
        
        if b_data and b_data['vals']:
            avg_temp = round(sum(b_data['vals']) / len(b_data['vals']), 1)
            min_temp = round(min(b_data['vals']), 1)
            max_temp = round(max(b_data['vals']), 1)
            all_avg_temps.append(avg_temp)
            
            is_high = (hi is not None and avg_temp > hi)
            is_low = (lo is not None and avg_temp < lo)
            if is_high:
                status_str = f"高溫超標 (>{hi}°C)"
                status_type = "ALARM_HIGH"
                out_of_bounds_count += 1
            elif is_low:
                status_str = f"低溫超標 (<{lo}°C)"
                status_type = "ALARM_LOW"
                out_of_bounds_count += 1
            else:
                status_str = "正常"
                status_type = "NORMAL"
                
            records.append({
                'time': b_key,
                'avg_temp': avg_temp,
                'min_temp': min_temp,
                'max_temp': max_temp,
                'status': status_str,
                'status_type': status_type
            })
        curr += timedelta(minutes=interval_min)

    total_samples = len(records)
    if all_avg_temps:
        max_t = max(all_avg_temps)
        min_t = min(all_avg_temps)
        mean_t = round(sum(all_avg_temps) / len(all_avg_temps), 1)
        compliance_rate = round(((total_samples - out_of_bounds_count) / total_samples) * 100, 1) if total_samples > 0 else 100.0
    else:
        max_t, min_t, mean_t, compliance_rate = None, None, None, 100.0

    summary = {
        'target_name': room_name,
        'report_type_name': '冷鏈庫溫歷程紀錄表',
        'hi': hi,
        'lo': lo,
        'total_samples': total_samples,
        'max_temp': max_t,
        'min_temp': min_t,
        'mean_temp': mean_t,
        'compliance_rate': compliance_rate,
        'out_of_bounds_count': out_of_bounds_count,
        'out_of_bounds_minutes': out_of_bounds_count * interval_min
    }

    return summary, records

METER_EQUIPMENT_MAP = {
    'all': [
        ('ch01', '1F 冷凍A'),
        ('ch02', '1F 冷凍B'),
        ('ch03', '1F 冷凍C'),
        ('ch04', '1F 冷凍D'),
        ('ch05', '1F 冷凍E'),
        ('ch06', '1F 緩衝A'),
        ('ch07', '1F 碼頭A'),
        ('ch08', '3F 急速20HP'),
        ('ch09', '3F 急速10HP'),
        ('ch10', '3F 半成品A'),
        ('ch11', '3F 半成品B'),
        ('ch12', '3F 冷藏A')
    ],
    'ch13': [
        ('ch01', '1F 冷凍A'),
        ('ch02', '1F 冷凍B'),
        ('ch03', '1F 冷凍C'),
        ('ch04', '1F 冷凍D'),
        ('ch05', '1F 冷凍E'),
        ('ch06', '1F 緩衝A'),
        ('ch07', '1F 碼頭A')
    ],
    'ch14': [
        ('ch08', '3F 急速20HP'),
        ('ch09', '3F 急速10HP'),
        ('ch10', '3F 半成品A'),
        ('ch11', '3F 半成品B'),
        ('ch12', '3F 冷藏A')
    ]
}

def _generate_energy_report_data(meter_id, d_from, d_to, interval_min=5):
    m_def = REPORT_METER_DEFINITIONS.get(meter_id)
    if not m_def:
        raise ValueError("未知的電表代碼")
    
    meter_name = m_def['name']
    chs = m_def['channels']
    equip_list = METER_EQUIPMENT_MAP.get(meter_id, [])
    equip_chs = [ch for ch, _ in equip_list]
    
    chs_placeholder = ', '.join(['%s'] * len(chs))
    with get_pg() as conn:
        rows = conn.execute(f"""
            SELECT timestamp, channel, v, a, kw, pf, kwh
            FROM power_readings
            WHERE channel IN ({chs_placeholder}) AND timestamp >= %s AND timestamp <= %s
            ORDER BY timestamp ASC
        """, chs + [d_from, d_to]).fetchall()

        equip_rows = []
        if equip_chs:
            equip_placeholder = ', '.join(['%s'] * len(equip_chs))
            equip_rows = conn.execute(f"""
                SELECT timestamp, channel, runtime_hours, compressor_current
                FROM temperatures
                WHERE channel IN ({equip_placeholder}) AND timestamp >= %s AND timestamp <= %s
                ORDER BY timestamp ASC
            """, equip_chs + [d_from, d_to]).fetchall()

    from collections import defaultdict
    buckets = defaultdict(lambda: {
        'samples': defaultdict(lambda: {'kw_sum': 0.0, 'a_sum': 0.0, 'v_list': [], 'pf_list': []}),
        'kwh_by_ch': {},
        'equip_rt': {}
    })
    
    for r in rows:
        ts = r['timestamp']
        ts_tw = ts.astimezone(TZ_TW_APP) if isinstance(ts, datetime) else datetime.fromisoformat(str(ts).replace(' ', 'T')).replace(tzinfo=TZ_TW_APP)
        floored_min = (ts_tw.minute // interval_min) * interval_min
        b_key = ts_tw.replace(minute=floored_min, second=0, microsecond=0).strftime('%Y-%m-%d %H:%M')
        ts_key = ts_tw.strftime('%Y-%m-%d %H:%M:%S')
        
        sample = buckets[b_key]['samples'][ts_key]
        if r['kw'] is not None:
            sample['kw_sum'] += float(r['kw'])
        if r['a'] is not None:
            sample['a_sum'] += float(r['a'])
        if r['v'] is not None:
            sample['v_list'].append(float(r['v']))
        if r['pf'] is not None:
            sample['pf_list'].append(float(r['pf']))
        if r['kwh'] is not None:
            buckets[b_key]['kwh_by_ch'][r['channel']] = float(r['kwh'])

    for r in equip_rows:
        ts = r['timestamp']
        ts_tw = ts.astimezone(TZ_TW_APP) if isinstance(ts, datetime) else datetime.fromisoformat(str(ts).replace(' ', 'T')).replace(tzinfo=TZ_TW_APP)
        floored_min = (ts_tw.minute // interval_min) * interval_min
        b_key = ts_tw.replace(minute=floored_min, second=0, microsecond=0).strftime('%Y-%m-%d %H:%M')
        
        ch = r['channel']
        rt = float(r['runtime_hours'] or 0.0)
        buckets[b_key]['equip_rt'][ch] = rt

    current_rt_state = {ch: 0.0 for ch in equip_chs}
    records = []
    curr = d_from.replace(minute=(d_from.minute // interval_min) * interval_min, second=0, microsecond=0)
    end_curr = d_to.replace(second=0, microsecond=0)
    
    all_kws = []
    all_v = []
    all_a = []
    all_pf = []
    
    while curr <= end_curr:
        b_key = curr.strftime('%Y-%m-%d %H:%M')
        b_data = buckets.get(b_key)
        
        if b_data and (b_data['samples'] or b_data['kwh_by_ch'] or b_data['equip_rt']):
            samples = list(b_data['samples'].values())
            if samples:
                # 採樣總和除以採樣筆數 (取得該區間之真正平均用電量與電流)
                kw_avg = round(sum(s['kw_sum'] for s in samples) / len(samples), 1)
                a_avg = round(sum(s['a_sum'] for s in samples) / len(samples), 1)
                
                v_all = [v for s in samples for v in s['v_list']]
                v_avg = round(sum(v_all) / len(v_all), 1) if v_all else 0.0
                
                pf_all = [pf for s in samples for pf in s['pf_list']]
                pf_avg = round(sum(pf_all) / len(pf_all), 2) if pf_all else 0.95
            else:
                kw_avg = 0.0
                a_avg = 0.0
                v_avg = 0.0
                pf_avg = 0.95
            
            kwh_sum = round(sum(b_data['kwh_by_ch'].values()), 1) if b_data['kwh_by_ch'] else 0.0
            
            for ch in equip_chs:
                if ch in b_data['equip_rt']:
                    current_rt_state[ch] = b_data['equip_rt'][ch]

            total_equip_rt = round(sum(current_rt_state.values()), 2)

            all_kws.append(kw_avg)
            if v_avg > 0: all_v.append(v_avg)
            if a_avg > 0: all_a.append(a_avg)
            if pf_avg > 0: all_pf.append(pf_avg)
            
            t_type = get_tariff_type(curr)
            if t_type == 'peak':
                tariff_type = "尖峰"
            elif t_type == 'semi_peak':
                tariff_type = "半尖峰"
            else:
                tariff_type = "離峰"
            
            records.append({
                'time': b_key,
                'tariff_type': tariff_type,
                'kw': kw_avg,
                'v': v_avg,
                'a': a_avg,
                'pf': pf_avg,
                'kwh': kwh_sum,
                'total_equip_runtime': total_equip_rt
            })
        curr += timedelta(minutes=interval_min)

    total_samples = len(records)
    avg_kw = round(sum(all_kws) / len(all_kws), 1) if all_kws else 0.0
    max_kw = max(all_kws) if all_kws else 0.0
    avg_v = round(sum(all_v) / len(all_v), 1) if all_v else 0.0
    avg_a = round(sum(all_a) / len(all_a), 1) if all_a else 0.0
    avg_pf = round(sum(all_pf) / len(all_pf), 2) if all_pf else 0.0
    
    first_kwh = records[0]['kwh'] if records else 0.0
    for idx_r, rec in enumerate(records):
        curr_kwh = rec['kwh']
        if idx_r == 0:
            rec['delta_kwh'] = 0.0
            rec['accum_kwh'] = 0.0
        else:
            prev_kwh = records[idx_r - 1]['kwh']
            delta = round(curr_kwh - prev_kwh, 1) if curr_kwh >= prev_kwh else 0.0
            accum = round(curr_kwh - first_kwh, 1) if curr_kwh >= first_kwh else 0.0
            rec['delta_kwh'] = delta
            rec['accum_kwh'] = accum

    last_accum = records[-1]['accum_kwh'] if records else 0.0
    consumed_kwh = round(last_accum, 1)

    summary = {
        'target_name': meter_name,
        'report_type_name': 'SPM-3 電表能源統計報表',
        'total_samples': total_samples,
        'consumed_kwh': consumed_kwh,
        'avg_kw': avg_kw,
        'max_kw': max_kw,
        'avg_v': avg_v,
        'avg_a': avg_a,
        'avg_pf': avg_pf,
        'total_equip_runtime_latest': round(sum(current_rt_state.values()), 2)
    }

    return summary, records

@app.route('/api/reports/query', methods=['GET'])
def query_report_data():
    try:
        category = request.args.get('category', 'temp')
        target = request.args.get('target', 'room1')
        period = request.args.get('period', 'day')
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        interval_min = int(request.args.get('interval_min', 5))

        d_from, d_to, label, _ = _calculate_report_time_range(period, date_from_str, date_to_str)

        if category == 'temp':
            summary, records = _generate_temperature_report_data(target, d_from, d_to, interval_min)
        else:
            summary, records = _generate_energy_report_data(target, d_from, d_to, interval_min)

        return jsonify({
            'status': 'ok',
            'category': category,
            'target': target,
            'period': period,
            'time_label': label,
            'interval_min': interval_min,
            'summary': summary,
            'records': records
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

@app.route('/api/reports/export', methods=['GET'])
def export_report_excel():
    try:
        category = request.args.get('category', 'temp')
        target = request.args.get('target', 'room1')
        period = request.args.get('period', 'day')
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        interval_min = int(request.args.get('interval_min', 5))

        d_from, d_to, label, filename_ts = _calculate_report_time_range(period, date_from_str, date_to_str)

        if category == 'temp':
            summary, records = _generate_temperature_report_data(target, d_from, d_to, interval_min)
            filename = f"冷鏈庫溫歷程紀錄表_{summary['target_name']}_{filename_ts}.xlsx"
            sheet_title = "庫溫紀錄表"
        else:
            summary, records = _generate_energy_report_data(target, d_from, d_to, interval_min)
            filename = f"SPM3電表能源統計報表_{summary['target_name']}_{filename_ts}.xlsx"
            sheet_title = "能源統計表"

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
        
        max_col = 6 if category == 'temp' else 9
        title_fill = PatternFill('solid', fgColor='0C3B82')
        hdr_fill = PatternFill('solid', fgColor='0C3B82')
        kpi_fill = PatternFill('solid', fgColor='F1F5F9')
        sub_fill = PatternFill('solid', fgColor='F1F5F9')
        alarm_fill = PatternFill('solid', fgColor='FEE2E2')
        thin_side = Side(style='thin', color='CBD5E1')
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        ws['A2'].value = f"統計區間：{label}   │   監控對象：{summary['target_name']}   │   產表時間：{datetime.now(TZ_TW_APP).strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(name='Arial', size=10, bold=True, color='475569')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

        if category == 'temp':
            meta_rows = [
                ("門檻基準", f"上限 {summary['hi']} °C / 下限 {summary['lo']} °C" if summary['hi'] is not None else "未設定", "合格率", f"{summary['compliance_rate']} %"),
                ("平均庫溫", f"{summary['mean_temp']} °C" if summary['mean_temp'] is not None else "--", "採樣筆數", f"{summary['total_samples']} 筆 (每{interval_min}分鐘)"),
                ("最高庫溫", f"{summary['max_temp']} °C" if summary['max_temp'] is not None else "--", "超標次數", f"{summary['out_of_bounds_count']} 次 (共{summary['out_of_bounds_minutes']}分鐘)"),
                ("最低庫溫", f"{summary['min_temp']} °C" if summary['min_temp'] is not None else "--", "備註說明", "客觀工程監控數據紀錄")
            ]
        else:
            meta_rows = [
                ("區間總耗電", f"{summary['consumed_kwh']} kWh", "平均功率因數", f"{summary['avg_pf']}"),
                ("即時用電量 (平均)", f"{summary['avg_kw']} kW", "採樣筆數", f"{summary['total_samples']} 筆 (每{interval_min}分鐘)"),
                ("最高用電量", f"{summary['max_kw']} kW", "平均線電壓", f"{summary['avg_v']} V"),
                ("總平均電流", f"{summary['avg_a']} A", "廠區設備總時數", f"{summary.get('total_equip_runtime_latest', 0.0)} hr")
            ]

        r_start = 4
        for r_offset, (k1, v1, k2, v2) in enumerate(meta_rows):
            row_num = r_start + r_offset
            for col_idx, text, is_label in [(1, k1, True), (2, v1, False), (4, k2, True), (5, v2, False)]:
                c = ws.cell(row_num, col_idx, text)
                c.border = cell_border
                c.font = Font(name='Arial', size=10, bold=is_label)
                if is_label:
                    c.fill = sub_fill
                    c.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    c.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
        max_col = 6 if category == 'temp' else 10

        # 主標題
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(1, 1, f"裕珍皇冷鏈監控系統 · {summary['report_type_name']}")
        title_cell.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 36

        # 報告摘要資訊
        meta_items = [
            ('監控對象：', summary['target_name'], '統計週期：', f"{label} ({period})"),
            ('採樣頻率：', f"每 {interval_min} 分鐘", '總資料筆數：', f"{summary['total_samples']} 筆"),
            ('產生時間：', datetime.now(TZ_TW_APP).strftime('%Y-%m-%d %H:%M:%S'), '報表系統：', 'YJH-SCADA Pro 2.0')
        ]
        for row_idx, (k1, v1, k2, v2) in enumerate(meta_items, 3):
            ws.cell(row_idx, 1, k1).font = Font(name='Arial', size=10, bold=True, color='475569')
            ws.cell(row_idx, 2, v1).font = Font(name='Arial', size=10, color='1E293B')
            ws.cell(row_idx, 4, k2).font = Font(name='Arial', size=10, bold=True, color='475569')
            ws.cell(row_idx, 5, v2).font = Font(name='Arial', size=10, color='1E293B')
            for col in range(1, max_col + 1):
                c = ws.cell(row_idx, col)
                c.fill = kpi_fill
                if col not in (1, 4):
                    c.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
            ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=min(6, max_col))

        # 數據表格表頭
        tbl_start_row = 9
        if category == 'temp':
            headers = ['#', '採樣時間', '平均庫溫 (°C)', '區間最低溫 (°C)', '區間最高溫 (°C)', '判定狀態']
            widths = [8, 20, 18, 18, 18, 22]
        else:
            headers = ['#', '採樣時間', '電費時段', '即時用電量 (平均, kW)', '平均線電壓 (V)', '總平均電流 (A)', '當期累計用電 (kWh)', '區間用電 (kWh)', '電表底度 (kWh)', '廠區設備運轉總時數 (hr)']
            widths = [8, 20, 14, 22, 18, 18, 22, 18, 20, 26]

        for c_idx, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(tbl_start_row, c_idx, h)
            cell.fill = hdr_fill
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cell_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w
        ws.row_dimensions[tbl_start_row].height = 24

        # 數據填充
        for idx, rec in enumerate(records, 1):
            r_num = tbl_start_row + idx
            ws.row_dimensions[r_num].height = 20
            is_alarm = (category == 'temp' and rec.get('status_type') != 'NORMAL')

            if category == 'temp':
                vals = [idx, rec['time'], rec['avg_temp'], rec['min_temp'], rec['max_temp'], rec['status']]
            else:
                vals = [idx, rec['time'], rec.get('tariff_type', '離峰'), rec['kw'], rec['v'], rec['a'], rec.get('accum_kwh', 0.0), rec.get('delta_kwh', 0.0), rec['kwh'], rec.get('total_equip_runtime', 0.0)]

            for c_idx, v in enumerate(vals, 1):
                cell = ws.cell(r_num, c_idx, v)
                cell.border = cell_border
                cell.font = Font(name='Arial', size=10)
                if is_alarm:
                    cell.fill = alarm_fill
                    if c_idx == 6:
                        cell.font = Font(name='Arial', size=10, bold=True, color='991B1B')

                if c_idx == 1 or c_idx == 2 or (category == 'energy' and c_idx == 3) or (category == 'temp' and c_idx == 6):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

            for c_idx, v in enumerate(vals, 1):
                cell = ws.cell(r_num, c_idx, v)
                cell.border = cell_border
                cell.font = Font(name='Arial', size=10)
                if is_alarm:
                    cell.fill = alarm_fill
                    if c_idx == 6:
                        cell.font = Font(name='Arial', size=10, bold=True, color='991B1B')

                if c_idx == 1 or c_idx == 2 or (category == 'temp' and c_idx == 6):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400
init_db()

MOCK_DATA_ENABLED = os.getenv('MOCK_DATA_ENABLED', 'false').strip().lower() in ('1', 'true', 'yes', 'on')

if MOCK_DATA_ENABLED:
    _start_mock_data_simulator()

if __name__ == '__main__':

    import sys

    if sys.platform == 'win32':

        from waitress import serve

        serve(app, host='0.0.0.0', port=88)

    else:

        app.run(debug=True, host='0.0.0.0', port=88)

