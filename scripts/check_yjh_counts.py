from openpyxl import load_workbook

path = r"C:\tmp\YJH_monitoring_site_template\inputs\裕珍皇_案場輸入總表.xlsx"
wb = load_workbook(path, data_only=True)
for sheet in ["01_案場", "02_庫別", "03_主機", "04_溫控器", "06_電表", "08_點位表", "10_告警"]:
    ws = wb[sheet]
    print(sheet, ws.max_row - 1)
