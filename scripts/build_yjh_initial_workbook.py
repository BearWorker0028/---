from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
INPUTS = ROOT / "inputs"
OUTPUTS = ROOT / "outputs"
DOCS = ROOT / "docs"

SITE_ID = "YJH001"

ROOMS = [
    ["R_1F_FREEZER", SITE_ID, "1F 冷凍庫", "冷凍", "待確認", "待確認", 5, "FP_1F", "1F 冷凍庫，對應 A-E 五台主機"],
    ["R_1F_BUFFER", SITE_ID, "1F 緩衝庫", "緩衝", "待確認", "待確認", 1, "FP_1F", "1F 緩衝庫，對應 A 主機"],
    ["R_1F_DOCK", SITE_ID, "1F 碼頭區", "碼頭", "待確認", "待確認", 1, "FP_1F", "1F 碼頭區，對應 A 主機"],
    ["R_3F_BLAST", SITE_ID, "3F 急速庫", "急速冷凍", "待確認", "待確認", 2, "FP_3F", "3F 急速庫，對應 A-B 二台主機"],
    ["R_3F_SEMI", SITE_ID, "3F 半成品庫", "冷藏/冷凍待確認", "待確認", "待確認", 2, "FP_3F", "3F 半成品庫，對應 A-B 二台主機"],
]

MACHINE_SPECS = [
    ("M_1F_FREEZER_A", "R_1F_FREEZER", "1F 冷凍庫 A"),
    ("M_1F_FREEZER_B", "R_1F_FREEZER", "1F 冷凍庫 B"),
    ("M_1F_FREEZER_C", "R_1F_FREEZER", "1F 冷凍庫 C"),
    ("M_1F_FREEZER_D", "R_1F_FREEZER", "1F 冷凍庫 D"),
    ("M_1F_FREEZER_E", "R_1F_FREEZER", "1F 冷凍庫 E"),
    ("M_1F_BUFFER_A", "R_1F_BUFFER", "1F 緩衝庫 A"),
    ("M_1F_DOCK_A", "R_1F_DOCK", "1F 碼頭區 A"),
    ("M_3F_BLAST_A", "R_3F_BLAST", "3F 急速庫 A"),
    ("M_3F_BLAST_B", "R_3F_BLAST", "3F 急速庫 B"),
    ("M_3F_SEMI_A", "R_3F_SEMI", "3F 半成品庫 A"),
    ("M_3F_SEMI_B", "R_3F_SEMI", "3F 半成品庫 B"),
]


def controller_id(machine_id: str) -> str:
    return "C_" + machine_id.removeprefix("M_")


def meter_id(name: str) -> str:
    return name.replace(" ", "_").replace("-", "_")


SHEETS = {
    "01_案場": {
        "headers": ["site_id", "site_name", "customer_name", "address", "timezone", "contact_name", "contact_phone", "maintenance_owner", "data_retention_days", "remarks"],
        "rows": [[SITE_ID, "裕珍皇", "裕珍皇", "高雄市茄萣區莒光路三段288-10號", "Asia/Taipei", "待確認", "待確認", "待確認", 365, "初版：由使用者提供之案場骨架建立"]],
    },
    "02_庫別": {
        "headers": ["room_id", "site_id", "room_name", "room_type", "target_temp_low_c", "target_temp_high_c", "machine_count", "floorplan_id", "remarks"],
        "rows": ROOMS,
    },
    "03_主機": {
        "headers": ["machine_id", "room_id", "machine_name", "machine_type", "brand", "model", "compressor_hp", "refrigerant", "controller_id", "meter_id", "enable_remote_control", "remarks"],
        "rows": [
            [machine_id, room_id, machine_name, "冷凍/冷藏主機待確認", "待確認", "待確認", "待確認", "待確認", controller_id(machine_id), "", "待確認", "主機規格待補"]
            for machine_id, room_id, machine_name in MACHINE_SPECS
        ],
    },
    "04_溫控器": {
        "headers": ["controller_id", "machine_id", "brand", "model", "protocol", "slave_id", "baud_rate", "parity", "data_bits", "stop_bits", "register_table_ref", "supports_write", "remarks"],
        "rows": [
            [controller_id(machine_id), machine_id, "IoT627", "待確認", "Modbus RTU", "待確認", "待確認", "待確認", 8, "待確認", "待提供", "待確認", "每台主機各 1 台 IoT627；站號/通訊參數待確認"]
            for machine_id, _room_id, _machine_name in MACHINE_SPECS
        ],
    },
    "05_偵測器": {
        "headers": ["sensor_id", "room_id", "machine_id", "sensor_type", "brand", "model", "signal_type", "protocol", "slave_id", "channel", "measure_min", "measure_max", "unit", "install_location", "calibration_offset", "remarks"],
        "rows": [],
    },
    "06_電表": {
        "headers": ["meter_id", "machine_id", "room_id", "brand", "model", "phase_type", "protocol", "slave_id", "baud_rate", "ct_ratio", "pt_ratio", "voltage_range", "current_range", "points_required", "remarks"],
        "rows": [
            ["PM_1F_CP_1", "", "", "待確認", "待確認", "待確認", "Modbus RTU", "待確認", "待確認", "待確認", "待確認", "待確認", "待確認", "V,A,kW,kWh,PF,Hz 待確認", "數位電表：1F CP-1，對應迴路/主機待確認"],
            ["PM_3F_CP_3", "", "", "待確認", "待確認", "待確認", "Modbus RTU", "待確認", "待確認", "待確認", "待確認", "待確認", "待確認", "V,A,kW,kWh,PF,Hz 待確認", "數位電表：3F CP-3，對應迴路/主機待確認"],
        ],
    },
    "07_Gateway_MQTT": {
        "headers": ["gateway_id", "site_id", "gateway_type", "mac_address", "gw_id", "local_ip", "mqtt_host", "mqtt_port", "tls_enabled", "tls_version", "client_id", "publish_topics", "subscribe_topics", "username", "password_ref", "remarks"],
        "rows": [["GW_YJH_01", SITE_ID, "待確認", "待確認", "待確認", "待確認", "待確認", "8801/8883 待確認", "待確認", "TLSV1_2 待確認", "YJH001_GW_01", "v1/rawdata;v1/realtime 待確認", "cmd/<device_id> 待確認", "", "待確認", "Gateway/MQTT 資訊尚未提供"]],
    },
    "08_點位表": {
        "headers": ["point_id", "site_id", "room_id", "machine_id", "device_id", "source_type", "protocol_key", "point_name", "point_type", "modbus_address", "bit_index", "data_type", "scale", "offset", "unit", "read_write", "polling_sec", "value_map", "normal_min", "normal_max", "display_on_dashboard", "record_history", "remarks"],
        "rows": [
            [f"P_{machine_id.removeprefix('M_')}_CTRL_TEMP", SITE_ID, room_id, machine_id, controller_id(machine_id), "controller", "待點位表確認", "控制溫度", "temperature", "待確認", "", "待確認", "待確認", 0, "degC", "R", 30, "", "待確認", "待確認", "Y", "Y", "由 IoT627 點位表補齊"]
            for machine_id, room_id, _machine_name in MACHINE_SPECS
        ],
    },
    "09_MQTT對照": {
        "headers": ["mqtt_key", "point_id", "direction", "confidence", "transform", "command_topic", "writable", "validation_note"],
        "rows": [],
    },
    "10_告警": {
        "headers": ["alarm_id", "point_id", "alarm_name", "severity", "condition_operator", "threshold", "duration_sec", "reset_condition", "notify_group_id", "enabled", "remarks"],
        "rows": [
            [f"A_{row[0]}_HIGH_TEMP", f"P_{row[0].removeprefix('M_')}_CTRL_TEMP", f"{row[2]} 高溫", "critical", ">", "待確認", "待確認", "待確認", "N_YJH_OPS", "Y", "溫度門檻待確認"]
            for row in MACHINE_SPECS
        ],
    },
    "11_通知": {
        "headers": ["notify_group_id", "channel", "recipient_name", "recipient_target", "schedule", "escalation_minutes", "enabled", "remarks"],
        "rows": [["N_YJH_OPS", "LINE/SMS/Email 待確認", "裕珍皇維護通知群組", "待確認", "24x7 待確認", "待確認", "Y", "通知對象待提供"]],
    },
    "12_平面圖": {
        "headers": ["floorplan_id", "file_name", "layer_name", "object_id", "object_type", "linked_id", "x", "y", "width", "height", "normal_style", "alarm_style", "remarks"],
        "rows": [
            ["FP_1F", "待提供_1F平面圖", "rooms", "OBJ_R_1F_FREEZER", "room", "R_1F_FREEZER", "待確認", "待確認", "待確認", "待確認", "green", "red", "待由平面圖標註"],
            ["FP_1F", "待提供_1F平面圖", "rooms", "OBJ_R_1F_BUFFER", "room", "R_1F_BUFFER", "待確認", "待確認", "待確認", "待確認", "green", "red", "待由平面圖標註"],
            ["FP_1F", "待提供_1F平面圖", "rooms", "OBJ_R_1F_DOCK", "room", "R_1F_DOCK", "待確認", "待確認", "待確認", "待確認", "green", "red", "待由平面圖標註"],
            ["FP_3F", "待提供_3F平面圖", "rooms", "OBJ_R_3F_BLAST", "room", "R_3F_BLAST", "待確認", "待確認", "待確認", "待確認", "green", "red", "待由平面圖標註"],
            ["FP_3F", "待提供_3F平面圖", "rooms", "OBJ_R_3F_SEMI", "room", "R_3F_SEMI", "待確認", "待確認", "待確認", "待確認", "green", "red", "待由平面圖標註"],
        ],
    },
    "13_Dashboard": {
        "headers": ["view_id", "view_name", "card_type", "linked_id", "point_id", "sort_order", "show_unit", "chart_enabled", "default_time_range", "role_visibility", "remarks"],
        "rows": [
            [f"V_{machine_id.removeprefix('M_')}_TEMP", f"{machine_name} 溫度卡", "point_card", machine_id, f"P_{machine_id.removeprefix('M_')}_CTRL_TEMP", i * 10, "Y", "Y", "24h", "admin;operator;viewer", "初版自動建立"]
            for i, (machine_id, _room_id, machine_name) in enumerate(MACHINE_SPECS, start=1)
        ],
    },
    "14_權限報表": {
        "headers": ["role_id", "role_name", "can_view", "can_control", "can_ack_alarm", "report_type", "report_schedule", "export_format", "retention_days", "remarks"],
        "rows": [["admin", "系統管理者", "Y", "Y", "Y", "daily_temperature", "daily 08:00 待確認", "xlsx;pdf", 365, ""], ["viewer", "客戶檢視", "Y", "N", "N", "monthly_summary", "monthly 待確認", "pdf", 365, ""]],
    },
    "15_檢核清單": {
        "headers": ["item_id", "category", "check_item", "owner", "required", "status", "evidence", "remarks"],
        "rows": [
            ["CHK_YJH_001", "基本資料", "確認庫別名稱與樓層正確", "工程師", "Y", "todo", "", ""],
            ["CHK_YJH_002", "溫控器", "確認 11 台 IoT627 型號、站號、通訊參數", "工程師", "Y", "todo", "", ""],
            ["CHK_YJH_003", "電表", "確認 1F CP-1、3F CP-3 品牌型號、CT/PT、站號", "工程師", "Y", "todo", "", ""],
            ["CHK_YJH_004", "平面圖", "提供 1F/3F 平面圖並標註庫別與設備位置", "工程師", "Y", "todo", "", ""],
            ["CHK_YJH_005", "Gateway", "確認 Gateway/MQTT IP、MAC、topic、TLS", "工程師", "Y", "todo", "", ""],
        ],
    },
}


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 12
        for cell in ws[letter]:
            max_len = max(max_len, min(36, len(str(cell.value or "")) + 2))
        ws.column_dimensions[letter].width = max_len


def build_workbook():
    INPUTS.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README")
    readme.append(["裕珍皇案場輸入總表", "", "", ""])
    readme.append(["建立時間", datetime.now().isoformat(timespec="seconds"), "", ""])
    readme.append(["目前資料狀態", "已填案場、庫別、主機、IoT627、電表骨架；其餘標示待確認", "", ""])
    readme.append(["下一步", "補溫控器站號/通訊參數、電表規格、Gateway/MQTT、點位表、平面圖", "", ""])
    style_sheet(readme)
    for name, spec in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(spec["headers"])
        for row in spec["rows"]:
            ws.append(row)
        style_sheet(ws)
    path = INPUTS / "裕珍皇_案場輸入總表.xlsx"
    wb.save(path)
    return path


def workbook_to_config(xlsx_path: Path):
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(xlsx_path, data_only=True)
    tables = {}
    for ws in wb.worksheets:
        if ws.title == "README":
            continue
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(v or "").strip() for v in rows[0]]
        records = []
        for values in rows[1:]:
            record = {headers[i]: values[i] if i < len(values) and values[i] is not None else "" for i in range(len(headers))}
            if any(str(v).strip() for v in record.values()):
                records.append(record)
        tables[ws.title] = records
    config = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "site_id": SITE_ID,
        "site_name": "裕珍皇",
        "status": "initial_skeleton",
        "tables": tables,
    }
    json_path = OUTPUTS / "site_config_initial.json"
    json_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    return json_path


def write_gap_report():
    lines = [
        "# 裕珍皇案場缺漏清單",
        "",
        "## 已建立",
        "",
        "- 案場：裕珍皇",
        "- 地址：高雄市茄萣區莒光路三段288-10號",
        "- 庫別：1F 冷凍庫、1F 緩衝庫、1F 碼頭區、3F 急速庫、3F 半成品庫",
        "- 主機：11 台",
        "- 溫控器：11 台 IoT627，每台主機 1 台",
        "- 數位電表：1F CP-1、3F CP-3",
        "",
        "## 待確認",
        "",
        "- 每個庫別的目標溫度範圍",
        "- 11 台 IoT627 的完整型號、RS485 站號、baud rate、parity、stop bit",
        "- IoT627 點位表或 Modbus register table",
        "- 主機品牌、型號、HP、冷媒",
        "- 數位電表品牌、型號、相別、CT/PT、站號、通訊參數、量測項目",
        "- Gateway 型號、MAC、IP、MQTT host/port/TLS/topic",
        "- 偵測器設備：門磁、溫濕度、壓力、電流、漏水等是否需要",
        "- 1F/3F 平面圖檔與設備座標",
        "- 告警門檻、延遲時間、通知對象",
    ]
    path = DOCS / "裕珍皇_缺漏清單.md"
    DOCS.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


if __name__ == "__main__":
    workbook = build_workbook()
    config = workbook_to_config(workbook)
    gaps = write_gap_report()
    print(workbook)
    print(config)
    print(gaps)
