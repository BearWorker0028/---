from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "outputs" / "site_inputs"


def slugify(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_]+", "_", value.strip())
    text = re.sub(r"_+", "_", text).strip("_")
    return text.upper() or "ITEM"


def parse_rooms(raw: str) -> list[dict]:
    rooms = []
    for idx, part in enumerate([p.strip() for p in raw.split(",") if p.strip()], start=1):
        pieces = [p.strip() for p in part.split(":")]
        name = pieces[0]
        machine_count = int(pieces[1]) if len(pieces) > 1 and pieces[1].isdigit() else 1
        room_type = pieces[2] if len(pieces) > 2 and pieces[2] else "待確認"
        room_id = f"R_{slugify(name)}" if re.search(r"[A-Za-z0-9]", name) else f"R_ROOM_{idx:02d}"
        rooms.append(
            {
                "room_id": room_id,
                "room_name": name,
                "room_type": room_type,
                "machine_count": machine_count,
                "floorplan_id": "FP_MAIN",
            }
        )
    if not rooms:
        rooms.append(
            {
                "room_id": "R_ROOM_01",
                "room_name": "庫別一",
                "room_type": "待確認",
                "machine_count": 1,
                "floorplan_id": "FP_MAIN",
            }
        )
    return rooms


def machine_rows(site_id: str, rooms: list[dict]) -> list[dict]:
    rows = []
    for room in rooms:
        prefix = room["room_id"].removeprefix("R_")
        for idx in range(1, int(room["machine_count"]) + 1):
            suffix = chr(64 + idx) if idx <= 26 else str(idx)
            machine_id = f"M_{prefix}_{suffix}"
            rows.append(
                {
                    "machine_id": machine_id,
                    "site_id": site_id,
                    "room_id": room["room_id"],
                    "machine_name": f"{room['room_name']} {suffix}",
                    "controller_id": f"C_{prefix}_{suffix}",
                    "meter_id": "",
                }
            )
    return rows


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 12
        for cell in ws[letter]:
            width = max(width, min(42, len(str(cell.value or "")) + 2))
        ws.column_dimensions[letter].width = width


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def build_workbook(site_id: str, site_name: str, customer_name: str, rooms: list[dict], output_dir: Path) -> Path:
    machines = machine_rows(site_id, rooms)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("README")
    readme.append(["新案場資料回填表", "", "", ""])
    readme.append(["建立時間", datetime.now().isoformat(timespec="seconds"), "", ""])
    readme.append(["使用方式", "先補 P0 欄位；未知資料填待確認，後續再回填。", "", ""])
    readme.append(["產出工具", "tools/site_input_builder.py", "", ""])
    style_sheet(readme)

    append_sheet(
        wb,
        "01_案場",
        [
            "site_id",
            "site_name",
            "customer_name",
            "address",
            "timezone",
            "contact_name",
            "contact_phone",
            "maintenance_owner",
            "remarks",
        ],
        [[site_id, site_name, customer_name, "待確認", "Asia/Taipei", "待確認", "待確認", "待確認", ""]],
    )

    append_sheet(
        wb,
        "02_庫別",
        [
            "room_id",
            "site_id",
            "room_name",
            "room_type",
            "target_temp_low_c",
            "target_temp_high_c",
            "machine_count",
            "floorplan_id",
            "remarks",
        ],
        [
            [
                room["room_id"],
                site_id,
                room["room_name"],
                room["room_type"],
                "待確認",
                "待確認",
                room["machine_count"],
                room["floorplan_id"],
                "",
            ]
            for room in rooms
        ],
    )

    append_sheet(
        wb,
        "03_主機",
        [
            "machine_id",
            "room_id",
            "machine_name",
            "machine_type",
            "brand",
            "model",
            "compressor_hp",
            "refrigerant",
            "controller_id",
            "meter_id",
            "enable_remote_control",
            "remarks",
        ],
        [
            [
                m["machine_id"],
                m["room_id"],
                m["machine_name"],
                "待確認",
                "待確認",
                "待確認",
                "待確認",
                "待確認",
                m["controller_id"],
                m["meter_id"],
                "待確認",
                "",
            ]
            for m in machines
        ],
    )

    append_sheet(
        wb,
        "04_Modbus通訊",
        [
            "device_id",
            "device_type",
            "linked_id",
            "protocol",
            "host",
            "port",
            "slave_id",
            "baud",
            "parity",
            "data_bits",
            "stop_bits",
            "remarks",
        ],
        [
            [
                m["controller_id"],
                "IoT627",
                m["machine_id"],
                "Modbus RTU",
                "待確認",
                "待確認",
                "待確認",
                "待確認",
                "待確認",
                8,
                "待確認",
                "",
            ]
            for m in machines
        ],
    )

    append_sheet(
        wb,
        "05_點位與警報",
        [
            "channel",
            "name",
            "linked_id",
            "slave_id",
            "register",
            "data_type",
            "scale",
            "offset",
            "unit",
            "hi",
            "lo",
            "delay_minutes",
            "alarm_enabled",
        ],
        [
            [
                f"ch{idx:02d}",
                room["room_name"],
                room["room_id"],
                "待確認",
                "待確認",
                "int16",
                0.1,
                0.0,
                "degC",
                "待確認",
                "",
                5,
                "Y",
            ]
            for idx, room in enumerate(rooms, start=1)
        ],
    )

    append_sheet(
        wb,
        "06_平面圖座標",
        ["floorplan_id", "file_name", "object_type", "linked_id", "display_name", "x", "y", "side", "remarks"],
        [
            [
                room["floorplan_id"],
                "待提供_floorplan.png",
                "room",
                room["room_id"],
                room["room_name"],
                "待確認",
                "待確認",
                "right",
                "",
            ]
            for room in rooms
        ],
    )

    append_sheet(
        wb,
        "07_報表欄位",
        ["report_key", "display_name", "unit", "enabled", "remarks"],
        [
            ["avg_temp", "平均庫溫", "degC", "Y", ""],
            ["temp_control", "控制溫度", "degC", "Y", ""],
            ["current", "運轉電流", "A", "Y", ""],
            ["kw", "即時耗電量", "kW", "Y", ""],
            ["kwh", "累積用電量", "kWh", "Y", ""],
        ],
    )

    path = output_dir / f"{site_id}_新案場資料回填表.xlsx"
    wb.save(path)
    return path


def workbook_to_config(workbook_path: Path, output_dir: Path) -> Path:
    wb = load_workbook(workbook_path, data_only=True)
    tables: dict[str, list[dict]] = {}
    for ws in wb.worksheets:
        if ws.title == "README":
            continue
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(v or "").strip() for v in rows[0]]
        records = []
        for values in rows[1:]:
            record = {
                headers[i]: values[i] if i < len(values) and values[i] is not None else ""
                for i in range(len(headers))
            }
            if any(str(value).strip() for value in record.values()):
                records.append(record)
        tables[ws.title] = records

    site = tables["01_案場"][0]
    config = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_workbook": str(workbook_path),
        "site_id": site["site_id"],
        "site_name": site["site_name"],
        "status": "skeleton",
        "tables": tables,
    }
    path = output_dir / f"{site['site_id']}_site_config.skeleton.json"
    path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def write_gap_report(site_id: str, site_name: str, rooms: list[dict], output_dir: Path) -> Path:
    lines = [
        f"# {site_name} 新案場缺漏清單",
        "",
        "## 已建立骨架",
        "",
        f"- site_id：{site_id}",
        f"- 案場名稱：{site_name}",
        f"- 庫別數：{len(rooms)}",
        f"- 預估主機數：{sum(int(room['machine_count']) for room in rooms)}",
        "",
        "## P0 待確認",
        "",
        "- Modbus 主機 IP / Port。",
        "- 每個通道 Slave ID、Register、資料型別、倍率。",
        "- 每個庫別高低溫警報門檻與延遲分鐘。",
        "- 實際平面圖與庫別座標。",
        "",
        "## P1 待確認",
        "",
        "- IoT627 / 電表 / DI 模組通訊參數。",
        "- 主機品牌、型號、HP、冷媒。",
        "- 報表欄位與顯示名稱。",
        "- 通知方式與通知對象。",
    ]
    path = output_dir / f"{site_id}_缺漏清單.md"
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description="建立通用新案場資料回填表與設定骨架。")
    parser.add_argument("--site-id", default="NEW_SITE", help="案場代碼，例如 YJH001")
    parser.add_argument("--site-name", default="新案場名稱", help="案場名稱")
    parser.add_argument("--customer-name", default="", help="客戶名稱；預設同案場名稱")
    parser.add_argument(
        "--rooms",
        default="庫別一:1:待確認",
        help="庫別清單，格式：名稱:主機數:類型,名稱:主機數:類型",
    )
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="輸出資料夾")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    site_id = slugify(args.site_id)
    site_name = args.site_name.strip() or "新案場名稱"
    customer_name = args.customer_name.strip() or site_name
    rooms = parse_rooms(args.rooms)

    workbook = build_workbook(site_id, site_name, customer_name, rooms, output_dir)
    config = workbook_to_config(workbook, output_dir)
    gaps = write_gap_report(site_id, site_name, rooms, output_dir)

    print(json.dumps({"workbook": str(workbook), "config": str(config), "gap_report": str(gaps)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
