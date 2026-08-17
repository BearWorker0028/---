# -*- coding: utf-8 -*-
"""
產生裕珍皇點位對照 Excel 檔 (包含 3 個工作頁表)
1. 頻道(卡片)對照表
2. iot627 實際點位表 (含遠端起停控制與 16 位元完整狀態 Bit Map)
3. spm-3實際點位表
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join("docs", "裕珍皇_監控點位與暫存器對照表.xlsx")

def create_excel():
    wb = Workbook()
    
    # 樣式設定
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    sub_header_font = Font(name="Microsoft JhengHei", size=10, bold=True, color="FFFFFF")
    sub_header_fill = PatternFill(start_color="2D6A9F", end_color="2D6A9F", fill_type="solid")

    title_font = Font(name="Microsoft JhengHei", size=14, bold=True, color="1F4E78")
    subtitle_font = Font(name="Microsoft JhengHei", size=10, italic=True, color="595959")
    
    cell_font = Font(name="Microsoft JhengHei", size=10)
    bold_font = Font(name="Microsoft JhengHei", size=10, bold=True)
    
    alt_fill = PatternFill(start_color="F2F7FA", end_color="F2F7FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    highlight_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # ---------------------------------------------------------
    # Sheet 1: 頻道(卡片)對照表
    # ---------------------------------------------------------
    ws1 = wb.active
    ws1.title = "頻道(卡片)對照表"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.append(["裕珍皇中央監控系統 - 頻道 (卡片) 對照表"])
    ws1.cell(row=1, column=1).font = title_font
    ws1.append(["說明：包含 1F / 3F 庫房點位、Gateway 網關與 Modbus Slave 通道對照"])
    ws1.cell(row=2, column=1).font = subtitle_font
    ws1.append([]) # 空行
    
    headers1 = [
        "通道 ID", "區域名稱 / 庫別", "通訊網關", "Modbus Slave ID", 
        "設備類型", "預設暫存器位址", "資料型別", "倍率", "單位", 
        "警報上限 (hi)", "警報下限 (lo)", "警報延遲", "狀態"
    ]
    ws1.append(headers1)
    
    data1 = [
        ["ch01", "1F 冷冷凍庫 A", "GW1 (1F 電盤)", 1, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", -15.0, None, "5 min", "啟用"],
        ["ch02", "1F 冷冷凍庫 B", "GW1 (1F 電盤)", 2, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", -15.0, None, "5 min", "啟用"],
        ["ch03", "1F 冷冷凍庫 C", "GW1 (1F 電盤)", 3, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", -15.0, None, "5 min", "啟用"],
        ["ch04", "1F 冷冷凍庫 D", "GW1 (1F 電盤)", 4, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", -15.0, None, "5 min", "啟用"],
        ["ch05", "1F 冷冷凍庫 E", "GW1 (1F 電盤)", 5, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", -15.0, None, "5 min", "啟用"],
        ["ch06", "1F 緩衝庫 A",   "GW1 (1F 電盤)", 6, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", 10.0,  None, "5 min", "啟用"],
        ["ch07", "1F 碼頭區 A",   "GW1 (1F 電盤)", 7, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", 15.0,  None, "5 min", "啟用"],
        ["ch08", "3F 急速庫 20HP", "GW2 (3F 電盤)", 13, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", -15.0, None, "5 min", "啟用"],
        ["ch09", "3F 急速庫 10HP", "GW2 (3F 電盤)", 14, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", -15.0, None, "5 min", "啟用"],
        ["ch10", "3F 半成品冷凍 A", "GW2 (3F 電盤)", 11, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", 8.0,   None, "5 min", "啟用"],
        ["ch11", "3F 半成品冷凍 B", "GW2 (3F 電盤)", 12, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", 8.0,   None, "5 min", "啟用"],
        ["ch12", "3F 冷藏庫",      "GW2 (3F 電盤)", 15, "IoT-627", "Offset 39 (40040)", "int16", 0.1, "°C", 8.0,   None, "5 min", "啟用"],
        ["ch13", "1F 集合式電錶", "GW1 (1F 電盤)", 8,  "SPM-3",   "Input Reg 1182",    "float32", 1.0, "kWh", None,  None, "5 min", "啟用"],
        ["ch14", "3F 集合式電錶", "GW2 (3F 電盤)", 16, "SPM-3",   "Input Reg 1182",    "float32", 1.0, "kWh", None,  None, "5 min", "啟用"],
    ]
    for r in data1:
        ws1.append(r)

    # ---------------------------------------------------------
    # Sheet 2: iot627 實際點位表 (含起停控制與 16 Bit Map 細項)
    # ---------------------------------------------------------
    ws2 = wb.create_sheet(title="iot627 實際點位表")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.append(["IoT-627 溫控器現場實際暫存器與起停狀態點位表"])
    ws2.cell(row=1, column=1).font = title_font
    ws2.append(["說明：Modbus Holding Registers (Function Code 03 / 06)，包含遠端起停控制、AI 監控點位與 16-Bit 狀態旗標細項"])
    ws2.cell(row=2, column=1).font = subtitle_font
    ws2.append([])
    
    headers2 = [
        "Offset (0-based)", "Modicon 40K 位址", "欄位名稱 (Field Name)", 
        "中文描述", "Function Code", "資料型別", "倍率 / 單位", "現場實測數值參考", "狀態說明與功能描述"
    ]
    ws2.append(headers2)
    
    data2_registers = [
        [0,  "40001", "power_on_off_cmd",        "設備遠端起停控制", "FC 03/06", "uint16", "開關狀態", "0: 停止, 1: 啟動", "主機遠端開關機控制指令 (0 = 關機/停止, 1 = 開機/啟動)"],
        [6,  "40007", "control_temperature_set", "設定控制溫度上限", "FC 03/06", "int16",  "0.1 °C",  "-18.0 °C", "溫控器設定控制溫度目標上限 (例如 -18.0°C)"],
        [7,  "40008", "temp_differential",       "控溫溫差/回差",   "FC 03/06", "int16",  "0.1 °C",  "3.0 °C",   "壓縮機啟停溫差/回差 (例如 3.0°C)"],
        [8,  "40009", "temp_offset",             "庫溫感溫頭校正",   "FC 03/06", "int16",  "0.1 °C",  "0.0 °C",   "庫內溫度感測器偏差補償校正值"],
        [34, "40035", "status_bitfield",          "設備完整狀態旗標", "FC 03",    "uint16", "16-Bit Map", "385 (0x0181)", "設備總體運轉/除霜/保護跳脫狀態 16-Bit 暫存器 (細項解析見下表)"],
        [39, "40040", "control_temperature",     "庫內控制溫度 AI", "FC 03",    "int16",  "0.1 °C",  "-18.2 °C", "庫內主要控制溫度感測點實測值"],
        [40, "40041", "coil_temperature",        "蒸發器盤管溫度 AI","FC 03",    "int16",  "0.1 °C",  "-21.0 °C", "蒸發器盤管感溫頭實測值 (除霜終止參考)"],
        [41, "40042", "return_pipe_temperature", "吸氣回流管溫度 AI","FC 03",    "int16",  "0.1 °C",  "-19.9 °C", "壓縮機吸氣管回溫感測點實測值"],
        [42, "40043", "low_pressure",            "系統低壓壓力 AI", "FC 03",    "int16",  "0.1 bar", "14.3 bar/psi", "壓縮機低壓側壓力感測實測值"],
        [44, "40045", "high_pressure",           "系統高壓壓力 AI", "FC 03",    "uint16", "0.1 bar", "249.4 bar/psi", "壓縮機高壓側壓力感測實測值"],
        [46, "40047", "compressor_current",      "壓縮機運轉電流 AI", "FC 03",    "int16",  "0.1 A",   "71.4 A",   "壓縮機三相運轉實測總電流"],
        [48, "40049", "defrost_current",         "除霜電熱電流 AI",  "FC 03",    "int16",  "0.1 A",   "0.1 A",    "除霜電熱管運轉實測電流"],
    ]
    for r in data2_registers:
        ws2.append(r)
        
    # 添加分隔空行與 Bitfield 細項子表格 Header
    ws2.append([])
    ws2.append(["IoT-627 Offset 34 (40035) 狀態 Bitfield 16 位元獨立對照細項"])
    sub_title_row = ws2.max_row
    ws2.cell(row=sub_title_row, column=1).font = bold_font
    
    headers2_bitfield = [
        "Bit 位元", "欄位名稱 (Field Name)", "中文狀態名稱", 
        "0 (OFF / 正常) 狀態", "1 (ON / 觸發) 狀態", "狀態類型", "對應控制/發報說明", "-", "-"
    ]
    ws2.append(headers2_bitfield)
    bitfield_header_row = ws2.max_row
    
    data2_bitmap = [
        ["Bit 0",  "running_status",   "設備起停運轉狀態", "0: 停止 (OFF)",  "1: 主機運轉中 (ON)", "起停狀態", "主機總體運轉/關機狀態指示 (面板起停切換視窗)"],
        ["Bit 1",  "defrost_status",   "除霜運轉狀態",     "0: 非除霜",       "1: 除霜運轉中 (ON)", "運轉狀態", "系統正在進行電熱/熱氣除霜動作"],
        ["Bit 2",  "drip_status",      "滴水延遲狀態",     "0: 非滴水",       "1: 滴水延遲中",     "延遲狀態", "除霜終止後之融冰水滴水防凍延遲"],
        ["Bit 3",  "fan_delay_status", "風機延遲啟動",     "0: 無延遲",       "1: 風機延遲中",     "延遲狀態", "防止除霜熱氣吹入庫內之風機啟動延遲"],
        ["Bit 4",  "high_temp_alarm",  "庫溫高溫警報",     "0: 正常",         "1: 高溫警報發報",   "溫度警報", "庫內溫度高於警報上限閥值 (自動觸發聲音/蜂鳴)"],
        ["Bit 5",  "low_temp_alarm",   "庫溫低溫警報",     "0: 正常",         "1: 低溫警報發報",   "溫度警報", "庫內溫度低於警報下限閥值"],
        ["Bit 6",  "defrost_heater",   "除霜電熱管狀態",   "0: 電熱關閉",     "1: 除霜電熱啟動",   "輸出狀態", "除霜電熱管加熱接觸器 (Heater Contactor) 輸出"],
        ["Bit 7",  "fan_status",       "蒸發風機運轉狀態", "0: 送風停止",     "1: 風機運轉中",     "輸出狀態", "庫內蒸發器送風機接觸器 (Fan Contactor) 輸出"],
        ["Bit 8",  "cooling_status",   "壓縮機製冷狀態",   "0: 非製冷",       "1: 壓縮機製冷中",   "輸出狀態", "壓縮機運轉電磁閥/電磁接觸器輸出"],
        ["Bit 9",  "low_press_err",    "低壓保護跳脫",     "0: 壓力正常",     "1: 低壓異常跳脫",   "硬體異常", "系統低壓壓力開關保護動作跳脫 (Low Press Error)"],
        ["Bit 10", "high_press_err",   "高壓保護跳脫",     "0: 壓力正常",     "1: 高壓異常跳脫",   "硬體異常", "系統高壓壓力開關保護動作跳脫 (High Press Error)"],
        ["Bit 11", "phase_err",        "電源逆相/缺相",    "0: 電源正常",     "1: 逆缺相異常跳脫", "硬體異常", "三相電源相序或缺相保護器動作"],
        ["Bit 12", "sensor_err",       "感溫頭故障",       "0: 感測正常",     "1: 感測器開路/短路", "硬體異常", "庫溫/盤管感溫頭訊號斷線或異常"],
        ["Bit 13", "overload_err",     "馬達積熱過載",     "0: 積熱正常",     "1: 過載跳脫異常",   "硬體異常", "壓縮機/風機過載積熱保護器動作"],
        ["Bit 14", "door_open",        "庫門開關狀態",     "0: 庫門關閉",     "1: 庫門開啟中",     "環境狀態", "庫門極限開關觸發 (開啟超過設定時間報警)"],
        ["Bit 15", "equip_err",        "設備綜合警報",     "0: 設備正常",     "1: 設備綜合故障",   "綜合警報", "任何硬體故障/異常保護動作導致之總警報跳脫"],
    ]
    for r in data2_bitmap:
        ws2.append(r)

    # ---------------------------------------------------------
    # Sheet 3: spm-3實際點位表
    # ---------------------------------------------------------
    ws3 = wb.create_sheet(title="spm-3實際點位表")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.append(["SPM-3 / S2-800MT 集合式電錶現場實際點位表"])
    ws3.cell(row=1, column=1).font = title_font
    ws3.append(["說明：Modbus Input Registers (Function Code 04)，IEEE 754 32-bit Float (Low-word first)"])
    ws3.cell(row=2, column=1).font = subtitle_font
    ws3.append([])
    
    headers3 = [
        "Input Register 位址", "Word 長度", "欄位名稱 (Field Name)", 
        "中文描述", "Function Code", "資料型別", "單位", "現場實測數值參考 (Slave #16)", "備註與說明"
    ]
    ws3.append(headers3)
    
    data3 = [
        ["1030 - 1031", 2, "voltage_a",        "相電壓 V_a",        "FC 04", "float32", "V",   "128.31 V",    "A 相對中性點相電壓"],
        ["1032 - 1033", 2, "voltage_rs",       "RS / AB 線電壓",     "FC 04", "float32", "V",   "223.73 V",    "RS 三相線電壓"],
        ["1034 - 1035", 2, "voltage_st",       "ST / BC 線電壓",     "FC 04", "float32", "V",   "221.99 V",    "ST 三相線電壓"],
        ["1036 - 1037", 2, "voltage_tr",       "TR / CA 線電壓",     "FC 04", "float32", "V",   "221.17 V",    "TR 三相線電壓"],
        ["1038 - 1039", 2, "voltage_ll_avg",   "三相平均線電壓",      "FC 04", "float32", "V",   "222.29 V",    "三相線電壓平均值"],
        ["1050 - 1051", 2, "frequency",        "電網頻率",          "FC 04", "float32", "Hz",  "59.98 Hz",    "系統頻率 (60Hz)"],
        ["1052 - 1053", 2, "current_r",        "R 相 / A 相電流",    "FC 04", "float32", "A",   "20.07 A",     "R 相即時電流"],
        ["1054 - 1055", 2, "current_s",        "S 相 / B 相電流",    "FC 04", "float32", "A",   "20.00 A",     "S 相即時電流"],
        ["1056 - 1057", 2, "current_t",        "T 相 / C 相電流",    "FC 04", "float32", "A",   "19.59 A",     "T 相即時電流"],
        ["1060 - 1061", 2, "current_avg",      "三相平均電流",        "FC 04", "float32", "A",   "12.47 A",     "三相平均電流"],
        ["1066 - 1067", 2, "current_sum",      "總有效電流",         "FC 04", "float32", "A",   "37.18 A",     "三相總電流"],
        ["1068 - 1069", 2, "power_r",          "R 相有效功率",       "FC 04", "float32", "kW",  "23.63 kW",    "R 相主功率"],
        ["1070 - 1071", 2, "power_s",          "S 相有效功率",       "FC 04", "float32", "kW",  "23.60 kW",    "S 相主功率"],
        ["1072 - 1073", 2, "power_t",          "T 相有效功率",       "FC 04", "float32", "kW",  "23.06 kW",    "T 相主功率"],
        ["1074 - 1075", 2, "power_total / kw", "廠區即時總有效功率", "FC 04", "float32", "kW",  "70.29 kW",    "廠區即時總用電功率"],
        ["1082 - 1083", 2, "power_factor / pf","總功率因數",         "FC 04", "float32", "-",   "0.85",        "系統總功率因數 (PF)"],
        ["1096 - 1097", 2, "reactive_power",   "總無功功率",         "FC 04", "float32", "kVAR","36.10 kVAR",  "總無功功率"],
        ["1098 - 1099", 2, "apparent_power",   "總視在功率",         "FC 04", "float32", "kVA", "52.43 kVA",   "總視在功率"],
        ["1102 - 1103", 2, "import_kwh",       "正向有功電量",       "FC 04", "float32", "kWh", "10,862.75 kWh","正向輸入電量"],
        ["1182 - 1183", 2, "energy_total / kwh","廠區總累積消耗電量", "FC 04", "float32", "kWh", "17,749.59 kWh","廠區總累積用電量 (主顯示)"],
    ]
    for r in data3:
        ws3.append(r)

    # ---------------------------------------------------------
    # 通用格式美化 (Headers, Borders, Auto-width)
    # ---------------------------------------------------------
    # Format Sheet 1 & Sheet 3 Headers
    for ws in [ws1, ws3]:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
            
        for row in range(5, ws.max_row + 1):
            is_alt = (row % 2 == 0)
            row_fill = alt_fill if is_alt else white_fill
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = align_center if col in [1, 3, 4, 5, 7, 8, 9, 10, 11, 12, 13] else align_left

    # Format Sheet 2 (Custom for 2 Tables inside)
    for col in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    for row in range(5, 17):
        is_alt = (row % 2 == 0)
        row_fill = alt_fill if is_alt else white_fill
        for col in range(1, ws2.max_column + 1):
            cell = ws2.cell(row=row, column=col)
            cell.font = cell_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = align_center if col in [1, 2, 5, 6, 7] else align_left
            
    # Format Sub-header for Bitfield Table
    for col in range(1, 10):
        cell = ws2.cell(row=bitfield_header_row, column=col)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    for row in range(bitfield_header_row + 1, ws2.max_row + 1):
        is_alt = (row % 2 == 0)
        row_fill = alt_fill if is_alt else white_fill
        for col in range(1, 10):
            cell = ws2.cell(row=row, column=col)
            cell.font = cell_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = align_center if col in [1, 4, 5, 6] else align_left

    # Set Row Heights & Column Widths
    for ws in [ws1, ws2, ws3]:
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[4].height = 28
        for r in range(5, ws.max_row + 1):
            if r not in [17, 18, 19]:
                ws.row_dimensions[r].height = 22
            
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                if length > max_len:
                    max_len = length
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    os.makedirs("docs", exist_ok=True)
    try:
        wb.save(OUTPUT_PATH)
        print(f"EXCEL CREATE SUCCESS: {OUTPUT_PATH}")
    except PermissionError:
        fallback_path = os.path.join("docs", "裕珍皇_監控點位與暫存器對照表_v2.xlsx")
        wb.save(fallback_path)
        print(f"PRIMARY LOCKED BY EXCEL. SAVED TO FALLBACK: {fallback_path}")

if __name__ == "__main__":
    create_excel()
